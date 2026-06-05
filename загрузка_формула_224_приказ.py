# INPUT: логика из загрузка_недели_амб_up.py (нормативы 224 приказа, витрина v_instrumental_examinations)
# OUTPUT: project/Расчет_загрузки_224_приказ.xlsx

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_PATH = Path(__file__).parent.parent / "Расчет_загрузки_224_приказ.xlsx"

# ── цвета ──────────────────────────────────────────────────────────────────
C_HEADER    = "1F4E79"   # тёмно-синий (белый текст)
C_SUBHEAD   = "2E75B6"   # синий
C_YELLOW    = "FFF2CC"   # светло-жёлтый (числитель)
C_GREEN     = "E2EFDA"   # светло-зелёный (знаменатель)
C_BLUE      = "DEEAF1"   # светло-голубой (результат)
C_ORANGE    = "FCE4D6"   # светло-оранжевый (промежуточный)
C_GRAY      = "F2F2F2"   # светло-серый
C_WHITE     = "FFFFFF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def border_all():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def border_thick():
    th = Side(style="medium", color="1F4E79")
    tn = Side(style="thin", color="BFBFBF")
    return Border(left=th, right=th, top=th, bottom=th)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header(cell, text, bg=C_HEADER):
    cell.value = text
    cell.font = font(bold=True, color="FFFFFF", size=10)
    cell.fill = fill(bg)
    cell.alignment = align("center")
    cell.border = border_all()

def style_subheader(cell, text, bg=C_SUBHEAD):
    cell.value = text
    cell.font = font(bold=True, color="FFFFFF", size=10)
    cell.fill = fill(bg)
    cell.alignment = align("center")
    cell.border = border_all()

def style_cell(cell, value, bg=C_WHITE, bold=False, h="center", num_fmt=None):
    cell.value = value
    cell.font = font(bold=bold)
    cell.fill = fill(bg)
    cell.alignment = align(h)
    cell.border = border_all()
    if num_fmt:
        cell.number_format = num_fmt

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ── ЛИСТ 1: Нормативы ──────────────────────────────────────────────────────
def sheet_normativy(wb):
    ws = wb.create_sheet("Нормативы (Пр.224)")
    ws.sheet_view.showGridLines = False

    # Заголовок
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "НОРМАТИВЫ ПО ПРИКАЗУ №224 — источник плановых значений для расчёта загрузки"
    c.font = font(bold=True, color="FFFFFF", size=12)
    c.fill = fill(C_HEADER)
    c.alignment = align("center")

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = "Данные нормативы жёстко прошиты в скрипт загрузка_недели_амб_up.py и применяются для расчёта плана"
    c.font = font(italic=True, size=9)
    c.fill = fill(C_GRAY)
    c.alignment = align("center")

    headers = [
        "Тип аппарата", "Тип пациентов", "Вид исследования",
        "Норматив времени, мин", "Кол-во в день (план)", "Мин в день (план)",
        "Откуда берётся?", "Что считается Фактом?", "Что считается Планом?",
        "Единица измерения", "Примечание"
    ]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i), h)

    rows = [
        ("КТ", "Взрослые (амб.)", "Натив (без КУ)",     "15",  "28", "420",
         "Приказ 224, Приложение 2",
         "кол-во DISTINCT accession_number с типом 'Натив'",
         "28 натив/день × рабочих_дней",
         "минуты", "нативных_исследований × 15"),
        ("КТ", "Взрослые (амб.)", "С КУ",               "25",  "11", "275",
         "Приказ 224, Приложение 2",
         "кол-во DISTINCT accession_number с типом 'С КУ'",
         "11 с_КУ/день × рабочих_дней",
         "минуты", "с_КУ_исследований × 25"),
        ("КТ", "Взрослые (амб.)", "ИТОГО день",          "—",   "39", "695",
         "—", "натив×15 + с_КУ×25", "(28×15 + 11×25) × рабочих_дней", "минуты", ""),
        ("КТ", "Дети (амб.)",    "Натив",                "20",  "19", "380",
         "Приказ 224",
         "кол-во DISTINCT accession_number с типом 'Натив'",
         "19 натив/день × рабочих_дней", "минуты", ""),
        ("КТ", "Дети (амб.)",    "С КУ",                 "35",  "8",  "280",
         "Приказ 224",
         "кол-во DISTINCT accession_number с типом 'С КУ'",
         "8 с_КУ/день × рабочих_дней", "минуты", ""),
        ("КТ", "Дети (амб.)",    "ИТОГО день",           "—",   "27", "660",
         "—", "натив×20 + с_КУ×35", "(19×20 + 8×35) × рабочих_дней", "минуты", ""),
        ("МРТ", "Взрослые (амб.)", "Натив",              "30",  "17", "510",
         "Приказ 224", "DISTINCT accession_number 'Натив'",
         "17 натив/день × рабочих_дней", "минуты", ""),
        ("МРТ", "Взрослые (амб.)", "С КУ",               "45",  "4",  "180",
         "Приказ 224", "DISTINCT accession_number 'С КУ'",
         "4 с_КУ/день × рабочих_дней", "минуты", ""),
        ("МРТ", "Взрослые (амб.)", "ИТОГО день",         "—",   "21", "690",
         "—", "натив×30 + с_КУ×45", "(17×30 + 4×45) × рабочих_дней", "минуты", ""),
        ("МРТ", "Дети (амб.)",   "Натив",                "40",  "13", "520",
         "Приказ 224", "DISTINCT accession_number 'Натив'",
         "13 натив/день × рабочих_дней", "минуты", ""),
        ("МРТ", "Дети (амб.)",   "С КУ",                 "60",  "2",  "120",
         "Приказ 224", "DISTINCT accession_number 'С КУ'",
         "2 с_КУ/день × рабочих_дней", "минуты", ""),
        ("МРТ", "Дети (амб.)",   "ИТОГО день",           "—",   "15", "640",
         "—", "натив×40 + с_КУ×60", "(13×40 + 2×60) × рабочих_дней", "минуты", ""),
        ("РГ",  "Взрослые (амб.)", "Исследования",       "—",   "72", "72",
         "Приказ 224", "кол-во DISTINCT accession_number", "72 × рабочих_дней",
         "штуки", "Факт = просто кол-во исследований"),
        ("ММГ", "Взрослые (амб.)", "Исследования",       "—",   "58", "58",
         "Приказ 224", "кол-во DISTINCT accession_number", "58 × рабочих_дней",
         "штуки", ""),
        ("Денситометрия", "Взрослые (амб.)", "Исследования", "—", "35", "35",
         "Приказ 224", "кол-во DISTINCT accession_number", "35 × рабочих_дней",
         "штуки", ""),
        ("Флюорография", "Взрослые", "Исследования",     "—",  "115","115",
         "Приказ 224", "кол-во DISTINCT accession_number", "115 × рабочих_дней",
         "штуки", ""),
        ("РГ",  "Дети (амб.)",   "Исследования",         "—",   "50", "50",
         "Приказ 224", "кол-во DISTINCT accession_number", "50 × рабочих_дней",
         "штуки", ""),
        ("Денситометрия", "Дети (амб.)", "Исследования", "—",   "24", "24",
         "Приказ 224", "кол-во DISTINCT accession_number", "24 × рабочих_дней",
         "штуки", ""),
    ]

    # Цвета по аппаратам
    color_map = {"КТ": "DEEAF1", "МРТ": "E2EFDA", "РГ": "FFF2CC",
                 "ММГ": "FCE4D6", "Денситометрия": "F8CBAD", "Флюорография": "F2F2F2"}

    for r, row in enumerate(rows, 4):
        bg = color_map.get(row[0], C_WHITE)
        for c_idx, val in enumerate(row, 1):
            style_cell(ws.cell(r, c_idx), val, bg=bg, h="left" if c_idx >= 7 else "center")

    widths = [18, 18, 20, 18, 18, 18, 28, 40, 40, 15, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A4"


# ── ЛИСТ 2: Методика ───────────────────────────────────────────────────────
def sheet_metodika(wb):
    ws = wb.create_sheet("Методика расчёта", 0)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 35

    def merge_title(row, text, bg=C_HEADER, colspan=5):
        ws.merge_cells(f"A{row}:{get_column_letter(colspan)}{row}")
        c = ws.cell(row, 1)
        c.value = text
        c.font = font(bold=True, color="FFFFFF", size=11)
        c.fill = fill(bg)
        c.alignment = align("center")
        ws.row_dimensions[row].height = 22

    def row_step(r, step, name, formula, source, bg=C_WHITE):
        ws.cell(r, 1).value = step
        ws.cell(r, 1).font = font(bold=True, size=12)
        ws.cell(r, 1).alignment = align("center")
        ws.cell(r, 1).fill = fill(bg)
        ws.cell(r, 1).border = border_all()
        style_cell(ws.cell(r, 2), name,    bg=bg, bold=True,  h="left")
        style_cell(ws.cell(r, 3), formula, bg=bg, bold=False, h="left")
        style_cell(ws.cell(r, 4), source,  bg=bg, bold=False, h="left")
        style_cell(ws.cell(r, 5), "",      bg=bg)
        ws.row_dimensions[r].height = 40

    merge_title(1, "АЛГОРИТМ РАСЧЁТА ПРОЦЕНТА ЗАГРУЗКИ ОБОРУДОВАНИЯ (Приказ №224)")
    merge_title(2,
        "Формула: Загрузка (%) = ФАКТ ÷ ПЛАН × 100    |    "
        "Источник данных: витрина data_views.v_instrumental_examinations (ЕРИС)",
        bg=C_SUBHEAD)

    ws.merge_cells("A3:E3")
    ws["A3"].value = ""

    # Шапка таблицы
    for col, txt in enumerate(["Шаг", "Что считаем", "Формула / значение", "Откуда берётся", "Примечание"], 1):
        style_header(ws.cell(4, col), txt)
    ws.row_dimensions[4].height = 20

    # --- КТ взрослые ---
    merge_title(5, "КТ (Компьютерная томография) — Взрослые амбулатория", bg="1A5276")
    row_step(6, "1", "Определяем аппарат",
             "ae_title → corrected_device_type = 'КТ'\n(по справочнику ae_title в скрипте)",
             "v_instrumental_examinations.ae_title\nМаппинг: AQSP_GP68, CT172868, REVOL_GP210...",
             C_GRAY)
    row_step(7, "2", "Считаем натив (числитель А)",
             "Натив = COUNT DISTINCT accession_number\nгде nativ_s_ku = 'Натив' (или NULL)",
             "v_instrumental_examinations.accession_number\n+ справочник процедур (nativ_s_ku)",
             C_YELLOW)
    row_step(8, "3", "Считаем С КУ (числитель Б)",
             "С_КУ = COUNT DISTINCT accession_number\nгде nativ_s_ku = 'С КУ'",
             "v_instrumental_examinations.accession_number\n+ справочник процедур (nativ_s_ku)",
             C_YELLOW)
    row_step(9, "4", "ФАКТ (числитель) — в минутах",
             "ФАКТ = Натив × 15 мин + С_КУ × 25 мин",
             "15 мин/натив, 25 мин/с_КУ — нормативы Приказа 224\n(Приложение 2, КТ взрослые)",
             C_YELLOW)
    row_step(10, "5", "Рабочих дней (знаменатель А)",
             "dni_po_normativu = кол-во раб. дней\nв данной неделе/месяце/дне",
             "Справочник normativy_dney_nedelya в скрипте\n(производственный календарь РФ)",
             C_GREEN)
    row_step(11, "6", "ПЛАН (знаменатель) — в минутах",
             "ПЛАН = (28 × 15 + 11 × 25) × рабочих_дней\n     = 695 × рабочих_дней",
             "28 натив/день + 11 с_КУ/день — норматив Приказа 224\nКТ взрослые: 39 исследований в день",
             C_GREEN)
    row_step(12, "7", "ЗАГРУЗКА (%)",
             "Загрузка = ФАКТ ÷ ПЛАН × 100\n(если ПЛАН = 0 → NULL)",
             "Итоговый показатель дашборда",
             C_BLUE)

    ws.merge_cells("A13:E13")
    ws["A13"].value = ""

    # --- МРТ ---
    merge_title(14, "МРТ (Магнитно-резонансная томография) — Взрослые амбулатория", bg="1A5276")
    row_step(15, "1-3", "Аппарат / Натив / С КУ",
             "Аналогично КТ (шаги 1-3)",
             "ae_title → МРТ: ING_GP45, AWP190202, EXCELMRI_GP219...", C_GRAY)
    row_step(16, "4", "ФАКТ (числитель)",
             "ФАКТ = Натив × 30 мин + С_КУ × 45 мин",
             "30 мин/натив, 45 мин/с_КУ — Приказ 224, МРТ взрослые", C_YELLOW)
    row_step(17, "5-6", "ПЛАН (знаменатель)",
             "ПЛАН = (17 × 30 + 4 × 45) × рабочих_дней\n     = 690 × рабочих_дней",
             "17 натив/день + 4 с_КУ/день (Приказ 224, МРТ взрослые)", C_GREEN)
    row_step(18, "7", "ЗАГРУЗКА (%)",
             "Загрузка = ФАКТ ÷ 690 ÷ рабочих_дней × 100", "—", C_BLUE)

    ws.merge_cells("A19:E19")
    ws["A19"].value = ""

    # --- ММГ, РГ, Денситометрия ---
    merge_title(20, "ММГ / РГ / Денситометрия — Взрослые амбулатория", bg="1A5276")
    row_step(21, "1", "Определяем аппарат", "ae_title → corrected_device_type",
             "ММГ: SENO_*, MAMMO* / РГ: RENEX*, RENEXRC* / Денситометрия: GELUNAR*", C_GRAY)
    row_step(22, "2", "ФАКТ (числитель)",
             "ФАКТ = COUNT DISTINCT accession_number\n(все исследования, без разбивки натив/с_КУ)",
             "v_instrumental_examinations.accession_number", C_YELLOW)
    row_step(23, "3", "ПЛАН (знаменатель)",
             "ММГ: 58 × рабочих_дней\n"
             "РГ:  72 × рабочих_дней\n"
             "Денситометрия: 35 × рабочих_дней",
             "Нормативы Приказа 224\n(кол-во исследований в день)", C_GREEN)
    row_step(24, "4", "ЗАГРУЗКА (%)",
             "Загрузка = ФАКТ ÷ ПЛАН × 100", "—", C_BLUE)

    ws.merge_cells("A25:E25")
    ws["A25"].value = ""
    merge_title(26,
        "ВАЖНО: для КТ и МРТ единица измерения — МИНУТЫ (время работы аппарата). "
        "Для РГ/ММГ/Денс — ШТУКИ (кол-во исследований).",
        bg="C00000")


# ── ЛИСТ 3: Март (месяц) ───────────────────────────────────────────────────
def sheet_month(wb):
    ws = wb.create_sheet("Март 2025 (месяц)")
    ws.sheet_view.showGridLines = False

    # Пример: ГП-219, 1 аппарат каждого типа, март 2025 = 21 рабочий день
    RABOCHY_DNEY = 21
    MO = "ГП №219 (пример)"
    PERIOD = "Март 2025"

    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = f"ПРИМЕР РАСЧЁТА ЗАГРУЗКИ ЗА МЕСЯЦ — {PERIOD} | {MO} | {RABOCHY_DNEY} рабочих дней"
    c.font = font(bold=True, color="FFFFFF", size=11)
    c.fill = fill(C_HEADER)
    c.alignment = align("center")

    ws.merge_cells("A2:P2")
    ws["A2"].value = ("Данные условные (пример). "
                      "Реальные данные берутся из: data_views.v_instrumental_examinations (ClickHouse / ЕРИС)")
    ws["A2"].font = font(italic=True, size=9)
    ws["A2"].fill = fill(C_GRAY)
    ws["A2"].alignment = align("center")

    headers1 = [
        "", "Аппарат", "МО", "Период",
        "── ОТКУДА БЕРЁТСЯ ──", "", "", "",
        "── ЧИСЛИТЕЛЬ (ФАКТ) ──", "", "",
        "── ЗНАМЕНАТЕЛЬ (ПЛАН) ──", "", "",
        "ИТОГ", ""
    ]
    headers2 = [
        "№", "Тип", "Учреждение", "Раб. дней",
        "Источник данных", "Кол-во натив\n(штук)", "Кол-во с КУ\n(штук)", "Всего\nисслед.",
        "Норматив\nвремени", "ФАКТ\n(мин / шт)", "Формула факта",
        "Норматив\n(план/день)", "ПЛАН\n(мин / шт)", "Формула плана",
        "Загрузка\n(%)", "Вывод"
    ]

    for i, h in enumerate(headers1, 1):
        c = ws.cell(3, i)
        if h.startswith("──"):
            c.value = h
            c.font = font(bold=True, color="FFFFFF", size=9)
            c.fill = fill(C_SUBHEAD)
            c.alignment = align("center")
            c.border = border_all()
        elif h:
            style_header(ws.cell(3, i), h)

    for i, h in enumerate(headers2, 1):
        style_header(ws.cell(4, i), h)
    ws.row_dimensions[4].height = 35

    # Данные (условные)
    data = [
        # (тип, нативн, с_ку, норм_нат, норм_ку, план_нат_день, план_ку_день, план_прочие, ед_изм)
        ("КТ",            450, 180, 15, 25, 28, 11, None, "мин"),
        ("МРТ",           300, 65,  30, 45, 17,  4, None, "мин"),
        ("ММГ",          1000,  0,   0,  0,  0,  0,   58, "шт"),
        ("РГ",           1300,  0,   0,  0,  0,  0,   72, "шт"),
        ("Денситометрия",  620,  0,   0,  0,  0,  0,   35, "шт"),
    ]

    color_rows = {
        "КТ": C_BLUE, "МРТ": C_GREEN, "ММГ": C_ORANGE,
        "РГ": C_YELLOW, "Денситометрия": "F8CBAD"
    }

    for r_off, (tip, nativ, sku, norm_n, norm_k, pln_n, pln_k, pln_proch, ed) in enumerate(data):
        row = 5 + r_off
        bg = color_rows[tip]

        if ed == "мин":
            fakt = nativ * norm_n + sku * norm_k
            plan = (pln_n * norm_n + pln_k * norm_k) * RABOCHY_DNEY
            formula_fakt = f"{nativ}×{norm_n} + {sku}×{norm_k} = {fakt}"
            formula_plan = f"({pln_n}×{norm_n} + {pln_k}×{norm_k})×{RABOCHY_DNEY} = {plan}"
            norm_str = f"{norm_n} мин/натив\n{norm_k} мин/с_КУ"
            plan_str = f"{pln_n} нат. + {pln_k} с_КУ в день"
            vsego = nativ + sku
        else:
            fakt = nativ
            plan = pln_proch * RABOCHY_DNEY
            formula_fakt = f"{fakt} исследований"
            formula_plan = f"{pln_proch}×{RABOCHY_DNEY} = {plan}"
            norm_str = "—"
            plan_str = f"{pln_proch} исследований в день"
            vsego = nativ

        zagruzka = round(fakt / plan * 100, 1) if plan else None
        vyvod = ("✅ В норме" if zagruzka and zagruzka >= 85
                 else ("⚠️ Ниже нормы" if zagruzka else "—"))

        cells = [
            r_off + 1, tip, MO, RABOCHY_DNEY,
            "v_instrumental_examinations\n(ЕРИС → ClickHouse)",
            nativ if ed == "мин" else "—",
            sku if ed == "мин" else "—",
            vsego,
            norm_str, fakt, formula_fakt,
            plan_str, plan, formula_plan,
            zagruzka, vyvod
        ]

        for c_idx, val in enumerate(cells, 1):
            cell = ws.cell(row, c_idx)
            cell.value = val
            cell.font = font(bold=(c_idx in (1, 10, 13, 15)), size=10)
            cell.fill = fill(bg)
            cell.alignment = align("center" if c_idx not in (5, 11, 12, 14) else "left")
            cell.border = border_all()
            if c_idx == 15 and val is not None:
                cell.number_format = '0.0"%"'
        ws.row_dimensions[row].height = 45

    # Ширины
    widths = [4, 16, 22, 10, 32, 14, 12, 10, 20, 12, 35, 25, 12, 35, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    # Пояснение внизу
    note_row = 11
    ws.merge_cells(f"A{note_row}:P{note_row}")
    c = ws.cell(note_row, 1)
    c.value = ("ПОЯСНЕНИЕ: КТ и МРТ считаются в МИНУТАХ рабочего времени аппарата (натив×норматив_мин + с_КУ×норматив_мин). "
               "РГ, ММГ, Денситометрия — в ШТУКАХ исследований. "
               "Рабочих дней берётся из производственного календаря РФ.")
    c.font = font(italic=True, size=9)
    c.fill = fill(C_GRAY)
    c.alignment = align("left")


# ── ЛИСТ 4: Неделя апрель ──────────────────────────────────────────────────
def sheet_week(wb):
    ws = wb.create_sheet("Неделя апрель 2025")
    ws.sheet_view.showGridLines = False

    RABOCHY_DNEY = 6   # неделя 15 апреля 2025 по производственному календарю
    MO = "ГП №219 (пример)"
    PERIOD = "Неделя 15 (7–12 апреля 2025)"
    NEDELYA = 15
    GOD = 2025

    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = f"ПРИМЕР РАСЧЁТА ЗАГРУЗКИ ЗА НЕДЕЛЮ — {PERIOD} | {MO} | {RABOCHY_DNEY} рабочих дней"
    c.font = font(bold=True, color="FFFFFF", size=11)
    c.fill = fill(C_HEADER)
    c.alignment = align("center")

    ws.merge_cells("A2:P2")
    ws["A2"].value = (f"Год: {GOD}, Неделя ISO: {NEDELYA}. "
                      "Рабочих дней = normativy_dney_nedelya WHERE god=2025 AND nedelya=15 → 6 дней")
    ws["A2"].font = font(italic=True, size=9)
    ws["A2"].fill = fill(C_GRAY)
    ws["A2"].alignment = align("center")

    # Строки расчёта — пошаговая таблица для каждого аппарата
    ws.merge_cells("A3:P3")
    ws["A3"].value = "ПОШАГОВЫЙ РАСЧЁТ"
    ws["A3"].font = font(bold=True, color="FFFFFF", size=10)
    ws["A3"].fill = fill(C_SUBHEAD)
    ws["A3"].alignment = align("center")

    headers = ["Шаг", "Параметр", "КТ", "МРТ", "ММГ", "РГ", "Денситометрия"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(4, i), h)

    data_steps = [
        ("1", "Период / неделя ISO", f"Неделя {NEDELYA} ({GOD})", f"Неделя {NEDELYA} ({GOD})",
         f"Неделя {NEDELYA} ({GOD})", f"Неделя {NEDELYA} ({GOD})", f"Неделя {NEDELYA} ({GOD})"),
        ("2", "Откуда берём рабочие дни",
         "normativy_dney_nedelya\n(справочник в скрипте)",
         "normativy_dney_nedelya",
         "normativy_dney_nedelya",
         "normativy_dney_nedelya",
         "normativy_dney_nedelya"),
        ("3", "Рабочих дней (знам. А)",
         str(RABOCHY_DNEY), str(RABOCHY_DNEY), str(RABOCHY_DNEY),
         str(RABOCHY_DNEY), str(RABOCHY_DNEY)),
        ("4", "Кол-во натив (числ. А)\nDISTINCT accession_number", "120", "80", "—", "—", "—"),
        ("5", "Кол-во с КУ (числ. Б)\nDISTINCT accession_number", "50", "18", "—", "—", "—"),
        ("6", "Всего исследований", "170", "98", "220", "360", "170"),
        ("7", "Норматив мин (натив / с_КУ)",
         "15 мин / 25 мин", "30 мин / 45 мин", "— (в шт.)", "— (в шт.)", "— (в шт.)"),
        ("8", "ФАКТ = ? (числитель)",
         "120×15 + 50×25 = 3 050 мин",
         "80×30 + 18×45 = 3 210 мин",
         "220 шт",
         "360 шт",
         "170 шт"),
        ("9", "Норматив план (в день)",
         "28 натив + 11 с_КУ → 695 мин", "17 натив + 4 с_КУ → 690 мин",
         "58 шт", "72 шт", "35 шт"),
        ("10", "ПЛАН = ? (знаменатель)",
         f"695 × {RABOCHY_DNEY} = {695*RABOCHY_DNEY} мин",
         f"690 × {RABOCHY_DNEY} = {690*RABOCHY_DNEY} мин",
         f"58 × {RABOCHY_DNEY} = {58*RABOCHY_DNEY} шт",
         f"72 × {RABOCHY_DNEY} = {72*RABOCHY_DNEY} шт",
         f"35 × {RABOCHY_DNEY} = {35*RABOCHY_DNEY} шт"),
        ("11", "ЗАГРУЗКА = ФАКТ÷ПЛАН×100",
         f"{round(3050/(695*RABOCHY_DNEY)*100,1)}%",
         f"{round(3210/(690*RABOCHY_DNEY)*100,1)}%",
         f"{round(220/(58*RABOCHY_DNEY)*100,1)}%",
         f"{round(360/(72*RABOCHY_DNEY)*100,1)}%",
         f"{round(170/(35*RABOCHY_DNEY)*100,1)}%"),
    ]

    colors_step = [C_GRAY, C_GRAY, C_GREEN, C_YELLOW, C_YELLOW, C_YELLOW,
                   C_ORANGE, C_YELLOW, C_GREEN, C_GREEN, C_BLUE]

    for r_off, (step, param, kt, mrt, mmg, rg, dens) in enumerate(data_steps):
        row = 5 + r_off
        bg = colors_step[r_off]
        vals = [step, param, kt, mrt, mmg, rg, dens]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row, c_idx)
            cell.value = val
            cell.font = font(bold=(r_off == 10), size=10)
            cell.fill = fill(bg)
            cell.alignment = align("left" if c_idx == 2 else "center")
            cell.border = border_all()
        ws.row_dimensions[row].height = 38

    widths = [5, 38, 28, 28, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


# ── ЛИСТ 5: День ───────────────────────────────────────────────────────────
def sheet_day(wb):
    ws = wb.create_sheet("День (14.04.2025)")
    ws.sheet_view.showGridLines = False

    RABOCHY_DNEY = 1
    MO = "ГП №219 (пример)"
    PERIOD = "14 апреля 2025"

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"ПРИМЕР РАСЧЁТА ЗАГРУЗКИ ЗА 1 ДЕНЬ — {PERIOD} | {MO}"
    c.font = font(bold=True, color="FFFFFF", size=11)
    c.fill = fill(C_HEADER)
    c.alignment = align("center")

    ws.merge_cells("A2:G2")
    ws["A2"].value = ("Рабочих дней = 1. "
                      "Данные за день: conduct_date = '2025-04-14' → фильтр в v_instrumental_examinations")
    ws["A2"].font = font(italic=True, size=9)
    ws["A2"].fill = fill(C_GRAY)
    ws["A2"].alignment = align("center")

    headers = ["Шаг", "Параметр", "КТ", "МРТ", "ММГ", "РГ", "Денситометрия"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i), h)

    data_steps = [
        ("1", "Дата", "14.04.2025", "14.04.2025", "14.04.2025", "14.04.2025", "14.04.2025"),
        ("2", "Источник данных",
         "v_instrumental_examinations\nWHERE conduct_date='2025-04-14'",
         "то же", "то же", "то же", "то же"),
        ("3", "Рабочих дней = 1", "1", "1", "1", "1", "1"),
        ("4", "Кол-во натив\n(DISTINCT accession_number)", "25", "15", "—", "—", "—"),
        ("5", "Кол-во с КУ\n(DISTINCT accession_number)", "10",  "4", "—", "—", "—"),
        ("6", "Всего исследований\n(DISTINCT accession_number)", "35", "19", "42", "68", "32"),
        ("7", "Норматив мин (натив/с_КУ)",
         "15 мин / 25 мин", "30 мин / 45 мин", "— (шт.)", "— (шт.)", "— (шт.)"),
        ("8", "ФАКТ (числитель)",
         "25×15 + 10×25 = 625 мин",
         "15×30 + 4×45 = 630 мин",
         "42 шт", "68 шт", "32 шт"),
        ("9", "Норм. план в день\n(знаменатель база)",
         "695 мин/день", "690 мин/день", "58 шт/день", "72 шт/день", "35 шт/день"),
        ("10", "ПЛАН = норматив × 1 день",
         "695 × 1 = 695 мин",
         "690 × 1 = 690 мин",
         "58 × 1 = 58 шт",
         "72 × 1 = 72 шт",
         "35 × 1 = 35 шт"),
        ("11", "ЗАГРУЗКА = ФАКТ÷ПЛАН×100",
         f"{round(625/695*100,1)}%",
         f"{round(630/690*100,1)}%",
         f"{round(42/58*100,1)}%",
         f"{round(68/72*100,1)}%",
         f"{round(32/35*100,1)}%"),
    ]

    colors_step = [C_GRAY, C_GRAY, C_GREEN, C_YELLOW, C_YELLOW, C_YELLOW,
                   C_ORANGE, C_YELLOW, C_GREEN, C_GREEN, C_BLUE]

    for r_off, row_data in enumerate(data_steps):
        row = 4 + r_off
        bg = colors_step[r_off]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row, c_idx)
            cell.value = val
            cell.font = font(bold=(r_off == 10), size=10)
            cell.fill = fill(bg)
            cell.alignment = align("left" if c_idx == 2 else "center")
            cell.border = border_all()
        ws.row_dimensions[row].height = 40

    widths = [5, 38, 30, 28, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── ЛИСТ 6: Сводная таблица ────────────────────────────────────────────────
def sheet_svodnaya(wb):
    ws = wb.create_sheet("Сводная таблица")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "СВОДНАЯ ТАБЛИЦА — расчёт загрузки по 5 типам аппаратов | ГП №219 (условный пример)"
    c.font = font(bold=True, color="FFFFFF", size=11)
    c.fill = fill(C_HEADER)
    c.alignment = align("center")

    headers = [
        "Аппарат", "Период",
        "Раб.\nдней",
        "Натив\n(факт, шт)",
        "С КУ\n(факт, шт)",
        "Норм.\nнатив (мин)",
        "Норм.\nс КУ (мин)",
        "ФАКТ\n(числитель)",
        "Ед.изм.\nфакта",
        "ПЛАН\n(знаменатель)",
        "Ед.изм.\nплана",
        "ЗАГРУЗКА\n(%)",
        "Статус"
    ]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(2, i), h)
    ws.row_dimensions[2].height = 35

    rows = [
        # Март
        ("КТ",   "Март 2025 (21 д.)",  21, 450, 180, 15, 25,
         450*15+180*25, "мин", 695*21, "мин", None),
        ("МРТ",  "Март 2025 (21 д.)",  21, 300,  65, 30, 45,
         300*30+65*45, "мин", 690*21, "мин", None),
        ("ММГ",  "Март 2025 (21 д.)",  21, 1000, 0, "—", "—",
         1000, "шт", 58*21, "шт", None),
        ("РГ",   "Март 2025 (21 д.)",  21, 1300, 0, "—", "—",
         1300, "шт", 72*21, "шт", None),
        ("Денс", "Март 2025 (21 д.)",  21,  620, 0, "—", "—",
         620, "шт", 35*21, "шт", None),
        # Неделя апрель
        ("КТ",   "Нед.15 апр. (6 д.)", 6, 120,  50, 15, 25,
         120*15+50*25, "мин", 695*6, "мин", None),
        ("МРТ",  "Нед.15 апр. (6 д.)", 6,  80,  18, 30, 45,
         80*30+18*45, "мин", 690*6, "мин", None),
        ("ММГ",  "Нед.15 апр. (6 д.)", 6, 220,   0, "—", "—",
         220, "шт", 58*6, "шт", None),
        ("РГ",   "Нед.15 апр. (6 д.)", 6, 360,   0, "—", "—",
         360, "шт", 72*6, "шт", None),
        ("Денс", "Нед.15 апр. (6 д.)", 6, 170,   0, "—", "—",
         170, "шт", 35*6, "шт", None),
        # День
        ("КТ",   "14.04.2025 (1 д.)",  1,  25,  10, 15, 25,
         25*15+10*25, "мин", 695*1, "мин", None),
        ("МРТ",  "14.04.2025 (1 д.)",  1,  15,   4, 30, 45,
         15*30+4*45, "мин", 690*1, "мин", None),
        ("ММГ",  "14.04.2025 (1 д.)",  1,  42,   0, "—", "—",
         42, "шт", 58*1, "шт", None),
        ("РГ",   "14.04.2025 (1 д.)",  1,  68,   0, "—", "—",
         68, "шт", 72*1, "шт", None),
        ("Денс", "14.04.2025 (1 д.)",  1,  32,   0, "—", "—",
         32, "шт", 35*1, "шт", None),
    ]

    color_by = {
        "КТ": C_BLUE, "МРТ": C_GREEN, "ММГ": C_ORANGE,
        "РГ": C_YELLOW, "Денс": "F8CBAD"
    }

    for r_off, row in enumerate(rows):
        r = 3 + r_off
        tip = row[0]
        fakt_val = row[7]
        plan_val = row[9]
        zagruzka = round(fakt_val / plan_val * 100, 1) if plan_val else None
        status = ("✅ Норма" if zagruzka and zagruzka >= 85
                  else ("⚠️ Ниже нормы" if zagruzka else "—"))
        bg = color_by.get(tip, C_WHITE)

        vals = list(row[:12]) + [zagruzka, status]
        # позиции: 1=тип,2=период,3=дней,4=натив,5=ску,6=норм_н,7=норм_к,8=факт,9=ед_ф,10=план,11=ед_п,12=загр,13=стат
        display = [row[0], row[1], row[2],
                   row[3] if row[3] != 0 or row[4] != 0 else "—",
                   row[4] if row[4] != 0 else "—",
                   row[5], row[6],
                   fakt_val, row[8],
                   plan_val, row[10],
                   zagruzka, status]

        for c_idx, val in enumerate(display, 1):
            cell = ws.cell(r, c_idx)
            cell.value = val
            cell.font = font(bold=(c_idx in (12, 13)), size=10)
            cell.fill = fill(bg)
            cell.alignment = align("center")
            cell.border = border_all()
            if c_idx == 12 and val is not None:
                cell.number_format = '0.0"%"'
        ws.row_dimensions[r].height = 22

    widths = [14, 22, 8, 14, 12, 14, 14, 14, 10, 14, 10, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # Легенда цветов
    legend_row = 20
    ws.merge_cells(f"A{legend_row}:M{legend_row}")
    ws.cell(legend_row, 1).value = "ЛЕГЕНДА:"
    ws.cell(legend_row, 1).font = font(bold=True)
    ws.cell(legend_row, 1).fill = fill(C_GRAY)
    ws.cell(legend_row, 1).alignment = align("left")

    legend = [
        (C_YELLOW,  "ФАКТ — числитель (то, что фактически выполнено)"),
        (C_GREEN,   "ПЛАН — знаменатель (норматив × рабочих дней)"),
        (C_BLUE,    "Итог загрузки = ФАКТ / ПЛАН × 100%"),
        (C_ORANGE,  "Промежуточные расчёты"),
    ]
    for i, (clr, txt) in enumerate(legend):
        ws.merge_cells(f"A{legend_row+1+i}:M{legend_row+1+i}")
        c = ws.cell(legend_row + 1 + i, 1)
        c.value = f"    {txt}"
        c.fill = fill(clr)
        c.font = font(size=9)
        c.alignment = align("left")
        c.border = border_all()


# ── MAIN ────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    # Удаляем дефолтный лист
    wb.remove(wb.active)

    sheet_metodika(wb)
    sheet_normativy(wb)
    sheet_month(wb)
    sheet_week(wb)
    sheet_day(wb)
    sheet_svodnaya(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"OK: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
