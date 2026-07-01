"""
minzdrav_equipment_hours_2026h1.py
Рабочие часы КТ и МРТ-аппаратов по неделям за I полугодие 2026 г. (для Минздрав).

Методология:
  1. conduct_date (DateTime) → toStartOfHour() → уникальные часовые слоты
  2. Для каждого аппарата × день — кол-во уникальных слотов = часов работы
  3. Сумма по дням внутри недели = часов за неделю
  4. Среднее по неделям полугодия = среднее часов в неделю

INPUT:  data_views.v_instrumental_examinations (source ClickHouse, через VPN)
OUTPUT: project/reports/minzdrav_kt_mrt_hours_2026h1.xlsx
"""

import subprocess
import pyautogui
import time
import sys
import os
from datetime import datetime

import clickhouse_connect
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import personal_config as cfg

# === Настройки ===
CH_HOST     = cfg.CH_HOST
CH_PORT     = cfg.CH_PORT
CH_USER     = cfg.CH_USER
CH_PASSWORD = cfg.CH_PASSWORD
CH_DATABASE = cfg.CH_DATABASE

VPN_APP_PATH            = cfg.VPN_APP_PATH
VPN_PASSWORD            = cfg.VPN_PASSWORD
PASSWORD_FIELD_X,  PASSWORD_FIELD_Y   = cfg.PASSWORD_FIELD_X,  cfg.PASSWORD_FIELD_Y
CONNECT_BUTTON_X,  CONNECT_BUTTON_Y   = cfg.CONNECT_BUTTON_X,  cfg.CONNECT_BUTTON_Y
RIGHT_CLICK_MENU_X, RIGHT_CLICK_MENU_Y = cfg.RIGHT_CLICK_MENU_X, cfg.RIGHT_CLICK_MENU_Y
DISCONNECT_MENU_ITEM_X, DISCONNECT_MENU_ITEM_Y = cfg.DISCONNECT_MENU_ITEM_X, cfg.DISCONNECT_MENU_ITEM_Y
CONFIRMATION_CLICK_X, CONFIRMATION_CLICK_Y = cfg.CONFIRMATION_CLICK_X, cfg.CONFIRMATION_CLICK_Y

DATE_FROM = '2026-01-01'
DATE_TO   = '2026-06-30'

OUTPUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'reports')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'minzdrav_kt_mrt_hours_2026h1.xlsx')


# === VPN ===

def connect_vpn():
    print("VPN: запуск...")
    try:
        subprocess.Popen(VPN_APP_PATH)
    except Exception as e:
        print(f"Ошибка запуска VPN: {e}")
        raise
    time.sleep(15)
    for window in pyautogui.getWindowsWithTitle(''):
        if any(kw in window.title.lower() for kw in ['check point', 'trgui', 'endpoint']):
            try:
                window.activate()
                time.sleep(2)
                break
            except Exception:
                pass
    pyautogui.click(PASSWORD_FIELD_X, PASSWORD_FIELD_Y)
    time.sleep(0.5)
    pyautogui.click(PASSWORD_FIELD_X, PASSWORD_FIELD_Y)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.3)
    pyautogui.press('delete')
    time.sleep(0.5)
    for char in VPN_PASSWORD:
        pyautogui.write(char)
        time.sleep(0.1)
    time.sleep(1)
    pyautogui.click(CONNECT_BUTTON_X, CONNECT_BUTTON_Y)
    time.sleep(15)
    print("VPN: подключён.")


def disconnect_vpn():
    print("VPN: отключение...")
    pyautogui.rightClick(RIGHT_CLICK_MENU_X, RIGHT_CLICK_MENU_Y)
    time.sleep(2)
    pyautogui.click(DISCONNECT_MENU_ITEM_X, DISCONNECT_MENU_ITEM_Y)
    time.sleep(3)
    pyautogui.click(CONFIRMATION_CLICK_X, CONFIRMATION_CLICK_Y)
    time.sleep(3)
    print("VPN: отключён.")


# === Запрос (только латинские алиасы) ===

QUERY = """
WITH daily AS (
    SELECT
        multiIf(
            vie.ae_title IN ('ING_GP45','AWP190202','AWP190378','AWP190384','AWP190464',
                             'AWP190802','AWP190826','EXCELMRI_GP219'),
            'МРТ',
            vie.ae_title IN ('AQSP_GP68','CT172868','REVOL_GP210','REVOL_GP214','RVMAX-GP23'),
            'КТ',
            vie.device_type
        ) AS tip_apparata,
        vie.ae_title,
        vie.conduct_mo_name,
        vie.conduct_mu_name,
        toYear(vie.conduct_date)    AS god,
        toWeek(vie.conduct_date, 3) AS nedelya,
        toString(toMonday(vie.conduct_date)) AS date_nedeli,
        toDate(vie.conduct_date)    AS den,
        uniqExact(toStartOfHour(vie.conduct_date)) AS chasov_v_den,
        dateDiff('hour', toStartOfHour(min(vie.conduct_date)), toStartOfHour(max(vie.conduct_date))) + 1 AS span_den
    FROM data_views.v_instrumental_examinations AS vie
    WHERE toDate(vie.conduct_date) >= '{date_from}'
      AND toDate(vie.conduct_date) <= '{date_to}'
      AND vie.ae_title IS NOT NULL
      AND vie.conduct_mo_name IS NOT NULL
      AND vie.conduct_date IS NOT NULL
    GROUP BY
        tip_apparata, vie.ae_title, vie.conduct_mo_name, vie.conduct_mu_name,
        god, nedelya, date_nedeli, den
    HAVING tip_apparata IN ('КТ', 'МРТ')
),
weekly AS (
    SELECT
        tip_apparata,
        ae_title,
        conduct_mo_name,
        conduct_mu_name,
        god,
        nedelya,
        date_nedeli,
        count()           AS dnei_otrabotano,
        sum(chasov_v_den) AS chasov_za_nedelyu,
        sum(span_den) AS span_za_nedelyu
    FROM daily
    GROUP BY
        tip_apparata, ae_title, conduct_mo_name, conduct_mu_name,
        god, nedelya, date_nedeli
)
SELECT
    tip_apparata,
    ae_title,
    conduct_mo_name,
    conduct_mu_name,
    god,
    nedelya,
    date_nedeli,
    dnei_otrabotano,
    chasov_za_nedelyu,
    round(
        avg(chasov_za_nedelyu) OVER (
            PARTITION BY tip_apparata, ae_title, conduct_mo_name, conduct_mu_name
        ), 1
    ) AS srednee_chasov,
    span_za_nedelyu,
    round(
        avg(span_za_nedelyu) OVER (
            PARTITION BY tip_apparata, ae_title, conduct_mo_name, conduct_mu_name
        ), 1
    ) AS srednee_span
FROM weekly
ORDER BY god, nedelya, tip_apparata, ae_title, conduct_mo_name
""".format(date_from=DATE_FROM, date_to=DATE_TO)

# Русские названия колонок для Excel
COLUMNS_RU = {
    'tip_apparata':      'Тип аппарата',
    'ae_title':          'AE Title',
    'conduct_mo_name':   'МО',
    'conduct_mu_name':   'Подразделение',
    'god':               'Год',
    'nedelya':           'Неделя',
    'date_nedeli':       'Дата начала недели',
    'dnei_otrabotano':   'Дней отработано',
    'chasov_za_nedelyu': 'Часов (слоты)',
    'srednee_chasov':    'Среднее (слоты)',
    'span_za_nedelyu':   'Часов (макс-мин)',
    'srednee_span':      'Среднее (макс-мин)',
}


# === Форматирование Excel ===

def _apply_header_style(ws):
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='CCCCCC')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[1].height = 36


def _apply_data_style(ws, n_rows: int):
    alt_fill = PatternFill(start_color='EEF3F9', end_color='EEF3F9', fill_type='solid')
    thin     = Side(style='thin', color='CCCCCC')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal='center', vertical='center')
    left     = Alignment(horizontal='left',   vertical='center')
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=n_rows + 1), start=1):
        for cell in row:
            cell.border    = border
            cell.font      = Font(size=10)
            cell.alignment = center if isinstance(cell.value, (int, float)) else left
            if i % 2 == 0:
                cell.fill = alt_fill


def format_xlsx(path: str, n_detail: int, n_summary: int):
    wb = load_workbook(path)

    ws = wb.active
    ws.title = 'По неделям'
    _apply_header_style(ws)
    _apply_data_style(ws, n_detail)
    for i, w in enumerate([12, 20, 42, 30, 6, 8, 16, 14, 16, 18, 16, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    ws2 = wb['Сводная']
    _apply_header_style(ws2)
    _apply_data_style(ws2, n_summary)
    for i, w in enumerate([12, 20, 42, 30, 10, 14, 16, 18, 16, 18], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    ws3 = wb.create_sheet('Методология')
    ws3.append(['Расчёт рабочих часов оборудования'])
    ws3.append([])
    ws3.append(['1. conduct_date (DateTime) → toStartOfHour() — округление до часа'])
    ws3.append(['2. Для каждого аппарата в каждый день — кол-во уникальных часовых слотов'])
    ws3.append(['   (= сколько разных часов хотя бы одно исследование было начато)'])
    ws3.append(['3. Сумма этих часов по всем дням недели = часов за неделю'])
    ws3.append(['4. Среднее (часов в неделю) = среднее по всем неделям полугодия'])
    ws3.append([])
    ws3.append(['Период:', f'{DATE_FROM} — {DATE_TO}'])
    ws3.append(['Источник:', 'data_views.v_instrumental_examinations'])
    ws3.append(['Сформировано:', datetime.now().strftime('%d.%m.%Y %H:%M')])
    ws3['A1'].font = Font(bold=True, size=12)

    wb.save(path)


# === Основная функция ===

def main():
    print(f"Запуск: рабочие часы КТ/МРТ за {DATE_FROM} — {DATE_TO}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    try:
        connect_vpn()
    except Exception as e:
        print(f"Ошибка VPN: {e}")
        sys.exit(1)

    client = None
    try:
        client = clickhouse_connect.get_client(
            host=CH_HOST, port=CH_PORT,
            username=CH_USER, password=CH_PASSWORD,
            database=CH_DATABASE,
            secure=True, verify=False,
            send_receive_timeout=94200, connect_timeout=999999
        )
        print("ClickHouse подключён. Выполняю запрос (может занять несколько минут)...")
        result = client.query(QUERY)
        client.close()
        client = None
    except Exception as e:
        if client:
            client.close()
        try:
            disconnect_vpn()
        except Exception:
            pass
        print(f"Ошибка ClickHouse: {e}")
        sys.exit(1)

    try:
        disconnect_vpn()
    except Exception:
        pass

    df = pd.DataFrame(result.result_rows, columns=list(COLUMNS_RU.keys()))
    df.rename(columns=COLUMNS_RU, inplace=True)
    n_detail = len(df)
    print(f"Получено строк (детализация): {n_detail}")

    if df.empty:
        print("Нет данных за указанный период.")
        sys.exit(0)

    # Сводная: среднее по аппарату за всё полугодие
    df_summary = (
        df.groupby(['Тип аппарата', 'AE Title', 'МО', 'Подразделение'], as_index=False)
        .agg(
            Недель_активности         = ('Неделя',              'nunique'),
            Дней_отработано           = ('Дней отработано',      'sum'),
            Итого_часов_слоты         = ('Часов (слоты)',        'sum'),
            Среднее_слоты             = ('Часов (слоты)',        'mean'),
            Итого_часов_макс_мин      = ('Часов (макс-мин)',     'sum'),
            Среднее_макс_мин          = ('Часов (макс-мин)',     'mean'),
        )
    )
    df_summary['Среднее_слоты']        = df_summary['Среднее_слоты'].round(1)
    df_summary['Итого_часов_макс_мин'] = df_summary['Итого_часов_макс_мин'].round(1)
    df_summary['Среднее_макс_мин']     = df_summary['Среднее_макс_мин'].round(1)
    n_summary = len(df_summary)

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='По неделям', index=False)
        df_summary.to_excel(writer, sheet_name='Сводная', index=False)

    format_xlsx(OUTPUT_FILE, n_detail, n_summary)

    print(f"Готово! {n_detail} строк детализации, {n_summary} аппаратов.")
    print(f"Файл: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
