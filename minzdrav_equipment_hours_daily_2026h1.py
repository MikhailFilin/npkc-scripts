"""
minzdrav_equipment_hours_daily_2026h1.py
Рабочие часы по всем аппаратам (ae_title) по дням за I полугодие 2026 г. (для Минздрав).

Методология (та же, что и в minzdrav_equipment_hours_2026h1.py, но без агрегации по неделям):
  1. conduct_date (DateTime) → toStartOfHour() → уникальные часовые слоты
  2. Для каждого аппарата × день:
     - chasov_slotov  = кол-во уникальных часовых слотов (сколько разных часов было исследование)
     - chasov_span    = (макс. час - мин. час в этот день) + 1

INPUT:  data_views.v_instrumental_examinations (source ClickHouse, через VPN)
OUTPUT: project/reports/minzdrav_kt_mrt_hours_daily_2026h1.xlsx
"""

import subprocess
import pyautogui
import time
import sys
import os
from datetime import datetime

import clickhouse_connect
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import personal_config as cfg

# === Настройки ===
CH_HOST     = cfg.CH_HOST
CH_PORT     = cfg.CH_PORT
CH_USER     = cfg.CH_USER
CH_PASSWORD = cfg.CH_PASSWORD
CH_DATABASE = cfg.CH_DATABASE

VPN_APP_PATH            = cfg.VPN_APP_PATH
VPN_PASSWORD            = cfg.VPN_PASSWORD
PASSWORD_FIELD_X,  PASSWORD_FIELD_Y   = cfg.PASSWORD_FIELD_X,  cfg.PASSWORD_FIELD_Y
CONNECT_BUTTON_X,  CONNECT_BUTTON_Y   = cfg.CONNECT_BUTTON_X,  cfg.CONNECT_BUTTON_Y
RIGHT_CLICK_MENU_X, RIGHT_CLICK_MENU_Y = cfg.RIGHT_CLICK_MENU_X, cfg.RIGHT_CLICK_MENU_Y
DISCONNECT_MENU_ITEM_X, DISCONNECT_MENU_ITEM_Y = cfg.DISCONNECT_MENU_ITEM_X, cfg.DISCONNECT_MENU_ITEM_Y
CONFIRMATION_CLICK_X, CONFIRMATION_CLICK_Y = cfg.CONFIRMATION_CLICK_X, cfg.CONFIRMATION_CLICK_Y

DATE_FROM = '2026-01-01'
DATE_TO   = '2026-06-30'

OUTPUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'reports')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'kometa_hours_daily_all_2026h1.xlsx')

INDEX_COLS = ['Тип аппарата', 'AE Title', 'МО', 'Подразделение']


# === VPN ===

def connect_vpn():
    print("VPN: запуск...")
    try:
        subprocess.Popen(VPN_APP_PATH)
    except Exception as e:
        print(f"Ошибка запуска VPN: {e}")
        raise
    time.sleep(15)
    for window in pyautogui.getWindowsWithTitle(''):
        if any(kw in window.title.lower() for kw in ['check point', 'trgui', 'endpoint']):
            try:
                window.activate()
                time.sleep(2)
                break
            except Exception:
                pass
    pyautogui.click(PASSWORD_FIELD_X, PASSWORD_FIELD_Y)
    time.sleep(0.5)
    pyautogui.click(PASSWORD_FIELD_X, PASSWORD_FIELD_Y)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.3)
    pyautogui.press('delete')
    time.sleep(0.5)
    for char in VPN_PASSWORD:
        pyautogui.write(char)
        time.sleep(0.1)
    time.sleep(1)
    pyautogui.click(CONNECT_BUTTON_X, CONNECT_BUTTON_Y)
    time.sleep(15)
    print("VPN: подключён.")


def disconnect_vpn():
    print("VPN: отключение...")
    pyautogui.rightClick(RIGHT_CLICK_MENU_X, RIGHT_CLICK_MENU_Y)
    time.sleep(2)
    pyautogui.click(DISCONNECT_MENU_ITEM_X, DISCONNECT_MENU_ITEM_Y)
    time.sleep(3)
    pyautogui.click(CONFIRMATION_CLICK_X, CONFIRMATION_CLICK_Y)
    time.sleep(3)
    print("VPN: отключён.")


# === Запрос (только латинские алиасы) ===

QUERY = """
SELECT
    any(tip_apparata)    AS tip_apparata,
    ae_title,
    any(conduct_mo_name) AS conduct_mo_name,
    any(conduct_mu_name) AS conduct_mu_name,
    den,
    uniqExact(hour_slot) AS chasov_slotov,
    dateDiff('hour', toStartOfHour(min(conduct_date)), toStartOfHour(max(conduct_date))) + 1 AS chasov_span
FROM (
    SELECT
        multiIf(
            vie.ae_title IN ('ING_GP45','AWP190202','AWP190378','AWP190384','AWP190464',
                             'AWP190802','AWP190826','EXCELMRI_GP219'),
            'МРТ',
            vie.ae_title IN ('AQSP_GP68','CT172868','REVOL_GP210','REVOL_GP214','RVMAX-GP23'),
            'КТ',
            vie.device_type
        ) AS tip_apparata,
        vie.ae_title            AS ae_title,
        vie.conduct_mo_name     AS conduct_mo_name,
        vie.conduct_mu_name     AS conduct_mu_name,
        toDate(vie.conduct_date) AS den,
        vie.conduct_date         AS conduct_date,
        toStartOfHour(vie.conduct_date) AS hour_slot
    FROM data_views.v_instrumental_examinations AS vie
    WHERE toDate(vie.conduct_date) >= '{date_from}'
      AND toDate(vie.conduct_date) <= '{date_to}'
      AND vie.ae_title IS NOT NULL
      AND vie.conduct_mo_name IS NOT NULL
      AND vie.conduct_date IS NOT NULL
) t
-- группируем строго по ae_title + день: у одного ae_title в разных строках
-- device_type (и, соответственно, tip_apparata) может отличаться — any() берёт
-- произвольное, но единое значение, чтобы не дробить один аппарат на несколько групп
GROUP BY ae_title, den
ORDER BY ae_title, den
""".format(date_from=DATE_FROM, date_to=DATE_TO)

# Русские названия колонок для Excel
COLUMNS_RU = {
    'tip_apparata':    'Тип аппарата',
    'ae_title':        'AE Title',
    'conduct_mo_name': 'МО',
    'conduct_mu_name': 'Подразделение',
    'den':             'Дата',
    'chasov_slotov':   'Часов (слоты)',
    'chasov_span':     'Часов (макс-мин)',
}


# === Форматирование Excel ===

def _apply_header_style(ws):
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='CCCCCC')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[1].height = 36


def format_xlsx(path: str, n_index_cols_matrix: int):
    wb = load_workbook(path)

    ws = wb['Детализация']
    _apply_header_style(ws)
    for i, w in enumerate([14, 20, 42, 30, 12, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    for sheet_name in ('Матрица часы-слоты', 'Матрица макс-мин'):
        ws2 = wb[sheet_name]
        _apply_header_style(ws2)
        for i in range(1, n_index_cols_matrix + 1):
            ws2.column_dimensions[get_column_letter(i)].width = [14, 20, 42, 30][i - 1]
        ws2.freeze_panes = get_column_letter(n_index_cols_matrix + 1) + '2'

    ws4 = wb['Сводная']
    _apply_header_style(ws4)
    for i, w in enumerate([14, 20, 42, 30, 14, 14, 16, 16, 18], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = 'A2'

    ws5 = wb.create_sheet('Методология')
    ws5.append(['Расчёт рабочих часов оборудования по дням (Комета, все AE Title)'])
    ws5.append([])
    ws5.append(['1. Источник: data_views.v_instrumental_examinations'])
    ws5.append(['2. conduct_date (DateTime) → toStartOfHour() — округление до часа'])
    ws5.append(['3. Группировка строго по ae_title + день (device_type может отличаться'])
    ws5.append(['   для одного ae_title — берётся произвольное единое значение)'])
    ws5.append(['4. "Часов (слоты)"    = кол-во уникальных часовых слотов за день'])
    ws5.append(['   (сколько разных часов хотя бы одно исследование было начато)'])
    ws5.append(['5. "Часов (макс-мин)" = (последний час - первый час за день) + 1'])
    ws5.append([])
    ws5.append(['Период:', f'{DATE_FROM} — {DATE_TO}'])
    ws5.append(['Источник:', 'data_views.v_instrumental_examinations'])
    ws5.append(['Сформировано:', datetime.now().strftime('%d.%m.%Y %H:%M')])
    ws5['A1'].font = Font(bold=True, size=12)

    wb.save(path)


# === Основная функция ===

def main():
    print(f"Запуск: рабочие часы всех аппаратов (Комета) по дням за {DATE_FROM} — {DATE_TO}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    try:
        connect_vpn()
    except Exception as e:
        print(f"Ошибка VPN: {e}")
        sys.exit(1)

    client = None
    try:
        client = clickhouse_connect.get_client(
            host=CH_HOST, port=CH_PORT,
            username=CH_USER, password=CH_PASSWORD,
            database=CH_DATABASE,
            secure=True, verify=False,
            send_receive_timeout=94200, connect_timeout=999999
        )
        print("ClickHouse подключён. Выполняю запрос (может занять несколько минут)...")
        result = client.query(QUERY)
        client.close()
        client = None
    except Exception as e:
        if client:
            client.close()
        try:
            disconnect_vpn()
        except Exception:
            pass
        print(f"Ошибка ClickHouse: {e}")
        sys.exit(1)

    try:
        disconnect_vpn()
    except Exception:
        pass

    df = pd.DataFrame(result.result_rows, columns=list(COLUMNS_RU.keys()))
    df.rename(columns=COLUMNS_RU, inplace=True)
    df['Дата'] = pd.to_datetime(df['Дата']).dt.date
    print(f"Получено строк (аппарат x день): {len(df)}, уникальных AE Title: {df['AE Title'].nunique()}")

    if df.empty:
        print("Нет данных за указанный период.")
        sys.exit(0)

    pivot_slots = df.pivot_table(
        index=INDEX_COLS, columns='Дата', values='Часов (слоты)', aggfunc='first'
    ).reset_index()
    pivot_slots.columns = [str(c) if not isinstance(c, str) else c for c in pivot_slots.columns]

    pivot_maxmin = df.pivot_table(
        index=INDEX_COLS, columns='Дата', values='Часов (макс-мин)', aggfunc='first'
    ).reset_index()
    pivot_maxmin.columns = [str(c) if not isinstance(c, str) else c for c in pivot_maxmin.columns]

    df_summary = (
        df.groupby(INDEX_COLS, as_index=False)
        .agg(
            Дней_активности   = ('Дата',              'nunique'),
            Итого_слоты       = ('Часов (слоты)',      'sum'),
            Среднее_слоты     = ('Часов (слоты)',      'mean'),
            Итого_макс_мин    = ('Часов (макс-мин)',   'sum'),
            Среднее_макс_мин  = ('Часов (макс-мин)',   'mean'),
        )
    )
    df_summary['Среднее_слоты']    = df_summary['Среднее_слоты'].round(1)
    df_summary['Среднее_макс_мин'] = df_summary['Среднее_макс_мин'].round(1)
    df_summary = df_summary.sort_values(INDEX_COLS)

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Детализация', index=False)
        pivot_slots.to_excel(writer, sheet_name='Матрица часы-слоты', index=False)
        pivot_maxmin.to_excel(writer, sheet_name='Матрица макс-мин', index=False)
        df_summary.to_excel(writer, sheet_name='Сводная', index=False)

    format_xlsx(OUTPUT_FILE, len(INDEX_COLS))

    print(f"Готово! {len(df)} строк детализации, {df_summary.shape[0]} аппаратов (AE Title).")
    print(f"Файл: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
