# INPUT: данные из витрины v_instrumental_examinations (вводятся вручную в Excel)
# OUTPUT: project/Калькулятор_загрузки.xlsx

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_PATH = Path(__file__).parent.parent / "Калькулятор_загрузки.xlsx"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def border_thick():
    t = Side(style="medium", color="1F4E79")
    return Border(left=t, right=t, top=t, bottom=t)

def font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, name="Calibri", italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr(ws, row, col, text, bg="1F4E79", fg="FFFFFF", bold=True, sz=10):
    c = ws.cell(row, col)
    c.value = text
    c.font = Font(bold=bold, size=sz, color=fg, name="Calibri")
    c.fill = fill(bg)
    c.alignment = align("center", wrap=True)
    c.border = border()
    return c

def inp(ws, row, col, value=None, note=None):
    """Ячейка ввода данных — жёлтая."""
    c = ws.cell(row, col)
    c.value = value
    c.font = font(bold=True, size=11)
    c.fill = fill("FFF2CC")
    c.alignment = align("center")
    c.border = border_thick()
    return c

def res(ws, row, col, formula, fmt=None, bg="DEEAF1"):
    """Ячейка с результатом — голубая, формула Excel."""
    c = ws.cell(row, col)
    c.value = formula
    c.font = font(bold=True, size=11)
    c.fill = fill(bg)
    c.alignment = align("center")
    c.border = border_thick()
    if fmt:
        c.number_format = fmt
    return c

def lbl(ws, row, col, text, bg="F2F2F2", bold=False, wrap=False, h="left"):
    c = ws.cell(row, col)
    c.value = text
    c.font = font(bold=bold, size=10)
    c.fill = fill(bg)
    c.alignment = align(h, wrap=wrap)
    c.border = border()
    return c

def w(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def merge(ws, r1, c1, r2, c2, text, bg="FFFFFF", bold=False, sz=10, fg="000000", h="left"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(r1, c1)
    c.value = text
    c.font = Font(bold=bold, size=sz, color=fg, name="Calibri")
    c.fill = fill(bg)
    c.alignment = align(h, wrap=True)
    c.border = border()
    return c


# ── Лист 1: НОРМАТИВЫ ──────────────────────────────────────────────────────
def make_normativy(wb):
    ws = wb.create_sheet("Нормативы")
    ws.sheet_view.showGridLines = False

    # Заголовок
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "НОРМАТИВЫ (Приказ №224)  —  используются в расчётах на листе Калькулятор"
    c.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    c.fill = fill("1F4E79")
    c.alignment = align("center")

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "Можно изменять значения в этой таблице — расчёты на листе Калькулятор пересчитаются автоматически"
    c.font = Font(bold=False, size=9, color="595959", italic=True, name="Calibri")
    c.fill = fill("EBF3FB")
    c.alignment = align("center")

    # Шапка
    for col, txt in enumerate(["Аппарат", "Тип пациентов", "Параметр", "Значение", "Единица", "Описание"], 1):
        hdr(ws, 3, col, txt)

    ws.row_dimensions[3].height = 22

    # Данные нормативов
    # Каждая строка: (аппарат, тип, параметр, значение, единица, описание, named_range_name)
    norms = [
        # КТ взрослые
        ("КТ", "Взрослые (амб.)", "Время натив",       15,  "мин/исслед.", "Длительность 1 нативного КТ"),
        ("КТ", "Взрослые (амб.)", "Время с КУ",        25,  "мин/исслед.", "Длительность 1 КТ с контрастом"),
        ("КТ", "Взрослые (амб.)", "Натив в день (план)", 28, "исслед./день", "Норматив кол-ва нативных за день"),
        ("КТ", "Взрослые (амб.)", "С КУ в день (план)",  11, "исслед./день", "Норматив кол-ва с КУ за день"),
        # МРТ взрослые
        ("МРТ", "Взрослые (амб.)", "Время натив",      30,  "мин/исслед.", "Длительность 1 нативного МРТ"),
        ("МРТ", "Взрослые (амб.)", "Время с КУ",       45,  "мин/исслед.", "Длительность 1 МРТ с контрастом"),
        ("МРТ", "Взрослые (амб.)", "Натив в день (план)", 17, "исслед./день", "Норматив нативных в день"),
        ("МРТ", "Взрослые (амб.)", "С КУ в день (план)",  4, "исслед./день", "Норматив с КУ в день"),
        # ММГ
        ("ММГ", "Взрослые (амб.)", "Исследований в день (план)", 58, "исслед./день", "Норматив для маммографа"),
        # РГ
        ("РГ",  "Взрослые (амб.)", "Исследований в день (план)", 72, "исслед./день", "Норматив для рентгена"),
        # Денситометрия
        ("Денситометрия", "Взрослые (амб.)", "Исследований в день (план)", 35, "исслед./день", "Норматив для денситометра"),
    ]

    colors = {"КТ": "DEEAF1", "МРТ": "E2EFDA", "ММГ": "FCE4D6", "РГ": "FFF2CC", "Денситометрия": "F8CBAD"}

    for r_off, (app, tip, param, val, ed, desc) in enumerate(norms):
        row = 4 + r_off
        bg = colors.get(app, "F2F2F2")
        for col, v in enumerate([app, tip, param, val, ed, desc], 1):
            c = ws.cell(row, col)
            c.value = v
            c.font = Font(bold=(col == 4), size=10, name="Calibri")
            c.fill = fill(bg)
            c.alignment = align("center" if col != 6 else "left")
            c.border = border()
        ws.row_dimensions[row].height = 20

    for col, w_val in zip("ABCDEF", [16, 20, 28, 12, 16, 42]):
        ws.column_dimensions[col].width = w_val

    ws.freeze_panes = "A4"
    return ws


# ── Лист 2: КАЛЬКУЛЯТОР ────────────────────────────────────────────────────
def make_kalkulator(wb):
    ws = wb.create_sheet("Калькулятор", 0)
    ws.sheet_view.showGridLines = False

    # Ширины столбцов
    widths = {"A": 2, "B": 24, "C": 18, "D": 18, "E": 18, "F": 2, "G": 22, "H": 16}
    for col, wid in widths.items():
        ws.column_dimensions[col].width = wid

    # ── Заголовок ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "КАЛЬКУЛЯТОР ЗАГРУЗКИ ОБОРУДОВАНИЯ (Приказ №224)"
    c.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    c.fill = fill("1F4E79")
    c.alignment = align("center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = ("Вводите данные в жёлтые ячейки  |  "
               "Нормативы — на листе «Нормативы»  |  "
               "Результат считается автоматически")
    c.font = Font(size=9, color="595959", italic=True, name="Calibri")
    c.fill = fill("EBF3FB")
    c.alignment = align("center")
    ws.row_dimensions[2].height = 18

    # Ссылки на нормативы (по номеру строки в листе Нормативы)
    # Нормативы!D4 = KT_nativ_min=15, D5=KT_sku_min=25, D6=KT_nativ_day=28, D7=KT_sku_day=11
    # D8=MRT_nativ_min=30, D9=MRT_sku_min=45, D10=MRT_nativ_day=17, D11=MRT_sku_day=4
    # D12=MMG_day=58, D13=RG_day=72, D14=DENS_day=35

    # ── Блок-легенда (мини) ──────────────────────────────────────────────────
    row = 4

    def section_header(r, title, bg="2E75B6"):
        ws.merge_cells(f"B{r}:H{r}")
        c = ws.cell(r, 2)
        c.value = title
        c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        c.fill = fill(bg)
        c.alignment = align("center")
        ws.row_dimensions[r].height = 20

    def col_headers(r):
        labels = ["Показатель", "Вводите данные\n(из витрины)", "", "",
                  "", "Нормативы\n(Пр.224)", "Результат\n(авто)"]
        bgs =    ["D9E1F2",     "FFF2CC",    "FFF2CC", "FFF2CC",
                  "FFFFFF",     "E2EFDA",    "DEEAF1"]
        for col_off, (lbl_txt, bg) in enumerate(zip(labels, bgs)):
            c = ws.cell(r, 2 + col_off)
            c.value = lbl_txt
            c.font = Font(bold=True, size=9, color="000000", name="Calibri")
            c.fill = fill(bg)
            c.alignment = align("center", wrap=True)
            c.border = border()
        ws.row_dimensions[r].height = 30

    def data_row(r, label, inp_col_b, inp_col_c, inp_col_d, norm_ref, result_formula,
                 inp_b_val=None, inp_c_val=None, inp_d_val=None,
                 bg_inp="FFF2CC", bg_res="DEEAF1", result_fmt='0.0"%"'):
        """Строка: [B]=label [C][D][E]=ввод [F]=пусто [G]=норматив [H]=результат."""
        lbl(ws, r, 2, label, bg="F7F7F7", bold=False, h="left")
        for col, val in zip([3, 4, 5], [inp_b_val, inp_c_val, inp_d_val]):
            inp(ws, r, col, val)
        ws.cell(r, 6).fill = fill("FFFFFF")
        # норматив
        nc = ws.cell(r, 7)
        nc.value = norm_ref
        nc.font = Font(bold=False, size=9, color="595959", italic=True, name="Calibri")
        nc.fill = fill("E2EFDA")
        nc.alignment = align("center", wrap=True)
        nc.border = border()
        # результат
        rc = ws.cell(r, 8)
        rc.value = result_formula
        rc.font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
        rc.fill = fill(bg_res)
        rc.alignment = align("center")
        rc.border = border_thick()
        if result_fmt:
            rc.number_format = result_fmt
        ws.row_dimensions[r].height = 22
        return rc

    # ══════════════════════════════════════════════════════════════════════════
    # ОБЩЕЕ: Рабочих дней
    # ══════════════════════════════════════════════════════════════════════════
    section_header(row, "ОБЩИЕ ВХОДНЫЕ ДАННЫЕ  (одинаковы для всех аппаратов)")
    row += 1

    ws.merge_cells(f"B{row}:E{row}")
    c = ws.cell(row, 2)
    c.value = "Рабочих дней в периоде (из производственного календаря / нормативного справочника)"
    c.font = Font(size=9, color="000000", name="Calibri")
    c.fill = fill("F7F7F7")
    c.alignment = align("left")
    c.border = border()
    ws.merge_cells(f"G{row}:H{row}")
    c2 = ws.cell(row, 7)
    c2.value = "← введите кол-во рабочих дней в жёлтую ячейку"
    c2.font = Font(size=9, italic=True, color="595959", name="Calibri")
    c2.fill = fill("FFF9C4")
    c2.alignment = align("center")
    c2.border = border()
    row += 1

    lbl(ws, row, 2, "Рабочих дней в периоде", bg="F7F7F7", bold=True, h="left")
    inp(ws, row, 3, 21)  # C
    ws.merge_cells(f"D{row}:E{row}")
    ws.cell(row, 4).value = "  ← вставьте сюда кол-во дней"
    ws.cell(row, 4).font = Font(size=9, italic=True, color="595959", name="Calibri")
    ws.cell(row, 4).fill = fill("FFFFF0")
    ws.cell(row, 4).alignment = align("left")
    ws.cell(row, 4).border = border()
    ws.cell(row, 6).fill = fill("FFFFFF")
    ws.cell(row, 7).fill = fill("FFFFFF")
    ws.cell(row, 8).fill = fill("FFFFFF")
    DAYS_CELL = f"C{row}"   # ← на эту ячейку ссылаются все формулы ниже
    ws.row_dimensions[row].height = 24
    row += 2

    # ══════════════════════════════════════════════════════════════════════════
    # КТ (взрослые амбулатория)
    # ══════════════════════════════════════════════════════════════════════════
    section_header(row, "КТ — Компьютерная томография  (взрослые, амбулатория)")
    row += 1
    for col, txt in zip([2,3,4,5,6,7,8],
                        ["Показатель", "Кол-во натив\n(из витрины)",
                         "Кол-во с КУ\n(из витрины)", "Всего\nисслед.", "",
                         "Норматив\n(Пр.224)", "= Результат"]):
        c = ws.cell(row, col)
        c.value = txt
        c.font = Font(bold=True, size=9, color="FFFFFF" if col == 2 else "000000", name="Calibri")
        c.fill = fill("2E75B6" if col == 2 else ("FFF2CC" if col in [3,4,5] else ("E2EFDA" if col==7 else ("DEEAF1" if col==8 else "FFFFFF"))))
        c.alignment = align("center", wrap=True)
        c.border = border()
    ws.row_dimensions[row].height = 30
    row += 1

    KT_NAT_ROW = row
    lbl(ws, row, 2, "Кол-во нативных КТ (натив)", "F7F7F7", h="left")
    inp(ws, row, 3, 450)   # C = натив
    inp(ws, row, 4, "")    # D (не нужен)
    ws.cell(row, 4).value = None
    ws.cell(row, 4).fill = fill("FAFAFA")
    ws.cell(row, 5).value = None; ws.cell(row, 5).fill = fill("FAFAFA"); ws.cell(row, 5).border = border()
    ws.cell(row, 6).fill = fill("FFFFFF")
    c = ws.cell(row, 7); c.value = "=Нормативы!D4"; c.font = Font(size=9, color="595959", italic=True, name="Calibri"); c.fill = fill("E2EFDA"); c.alignment = align("center"); c.border = border(); c.number_format = '0" мин/натив"'
    ws.cell(row, 8).fill = fill("FFFFFF"); ws.cell(row, 8).border = border()
    ws.row_dimensions[row].height = 20
    row += 1

    KT_SKU_ROW = row
    lbl(ws, row, 2, "Кол-во с контрастом (с КУ)", "F7F7F7", h="left")
    inp(ws, row, 3, 180)
    ws.cell(row, 4).value = None; ws.cell(row, 4).fill = fill("FAFAFA"); ws.cell(row, 4).border = border()
    ws.cell(row, 5).value = None; ws.cell(row, 5).fill = fill("FAFAFA"); ws.cell(row, 5).border = border()
    ws.cell(row, 6).fill = fill("FFFFFF")
    c = ws.cell(row, 7); c.value = "=Нормативы!D5"; c.font = Font(size=9, color="595959", italic=True, name="Calibri"); c.fill = fill("E2EFDA"); c.alignment = align("center"); c.border = border(); c.number_format = '0" мин/с_КУ"'
    ws.cell(row, 8).fill = fill("FFFFFF"); ws.cell(row, 8).border = border()
    ws.row_dimensions[row].height = 20
    row += 1

    KT_FAKT_ROW = row
    lbl(ws, row, 2, "ФАКТ (числитель)", "FFF2CC", bold=True, h="left")
    c = ws.cell(row, 3)
    c.value = f"=C{KT_NAT_ROW}*Нормативы!D4+C{KT_SKU_ROW}*Нормативы!D5"
    c.font = Font(bold=True, size=10, name="Calibri")
    c.fill = fill("FFEB9C"); c.alignment = align("center"); c.border = border_thick()
    c.number_format = '0" мин"'
    ws.merge_cells(f"D{row}:E{row}")
    c2 = ws.cell(row, 4)
    c2.value = f"=  натив × {chr(8203)}Норм_натив  +  с_КУ × Норм_с_КУ"
    c2.font = Font(size=9, italic=True, color="595959", name="Calibri")
    c2.fill = fill("FFEB9C"); c2.alignment = align("left"); c2.border = border()
    ws.cell(row, 6).fill = fill("FFFFFF")
    ws.cell(row, 7).value = "Нат×15 + С_КУ×25 мин"; ws.cell(row, 7).font = Font(size=9, color="595959", name="Calibri"); ws.cell(row, 7).fill = fill("E2EFDA"); ws.cell(row, 7).alignment = align("center"); ws.cell(row, 7).border = border()
    ws.cell(row, 8).fill = fill("FFFFFF"); ws.cell(row, 8).border = border()
    ws.row_dimensions[row].height = 22
    row += 1

    KT_PLAN_ROW = row
    lbl(ws, row, 2, "ПЛАН (знаменатель)", "E2EFDA", bold=True, h="left")
    c = ws.cell(row, 3)
    c.value = f"=(Нормативы!D6*Нормативы!D4+Нормативы!D7*Нормативы!D5)*{DAYS_CELL}"
    c.font = Font(bold=True, size=10, name="Calibri")
    c.fill = fill("C6EFCE"); c.alignment = align("center"); c.border = border_thick()
    c.number_format = '0" мин"'
    ws.merge_cells(f"D{row}:E{row}")
    c2 = ws.cell(row, 4)
    c2.value = "=  (28×15 + 11×25) × рабочих_дней  =  695 × дни"
    c2.font = Font(size=9, italic=True, color="595959", name="Calibri")
    c2.fill = fill("C6EFCE"); c2.alignment = align("left"); c2.border = border()
    ws.cell(row, 6).fill = fill("FFFFFF")
    ws.cell(row, 7).value = "695 мин/день × дней"; ws.cell(row, 7).font = Font(size=9, color="595959", name="Calibri"); ws.cell(row, 7).fill = fill("E2EFDA"); ws.cell(row, 7).alignment = align("center"); ws.cell(row, 7).border = border()
    ws.cell(row, 8).fill = fill("FFFFFF"); ws.cell(row, 8).border = border()
    ws.row_dimensions[row].height = 22
    row += 1

    lbl(ws, row, 2, "% ЗАГРУЗКИ КТ", "1F4E79", bold=True, h="left")
    ws.cell(row, 2).font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    c = ws.cell(row, 3)
    c.value = f"=IF(C{KT_PLAN_ROW}=0,\"\",C{KT_FAKT_ROW}/C{KT_PLAN_ROW}*100)"
    c.font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
    c.fill = fill("BDD7EE"); c.alignment = align("center"); c.border = border_thick()
    c.number_format = '0.0"%"'
    ws.merge_cells(f"D{row}:E{row}")
    ws.cell(row, 4).value = "= ФАКТ ÷ ПЛАН × 100"
    ws.cell(row, 4).font = Font(bold=True, size=10, italic=True, color="1F4E79", name="Calibri")
    ws.cell(row, 4).fill = fill("BDD7EE"); ws.cell(row, 4).alignment = align("left"); ws.cell(row, 4).border = border()
    ws.cell(row, 6).fill = fill("FFFFFF")
    # Статус
    sc = ws.cell(row, 7)
    sc.value = f'=IF(C{row}="","—",IF(C{row}>=85,"✓ Норма","! Ниже нормы"))'
    sc.font = Font(bold=True, size=10, name="Calibri"); sc.fill = fill("BDD7EE"); sc.alignment = align("center"); sc.border = border_thick()
    ws.cell(row, 8).fill = fill("FFFFFF"); ws.cell(row, 8).border = border()
    ws.row_dimensions[row].height = 28
    row += 2

    # ══════════════════════════════════════════════════════════════════════════
    # МРТ
    # ══════════════════════════════════════════════════════════════════════════
    section_header(row, "МРТ — Магнитно-резонансная томография  (взрослые, амбулатория)")
    row += 1
    for col, txt in zip([2,3,4,5,6,7,8], ["Показатель","Кол-во натив\n(из витрины)","Кол-во с КУ\n(из витрины)","Всего","","Норматив","= Результат"]):
        c = ws.cell(row, col); c.value = txt
        c.font = Font(bold=True, size=9, color="FFFFFF" if col==2 else "000000", name="Calibri")
        c.fill = fill("2E75B6" if col==2 else ("FFF2CC" if col in [3,4,5] else ("E2EFDA" if col==7 else ("DEEAF1" if col==8 else "FFFFFF"))))
        c.alignment = align("center", wrap=True); c.border = border()
    ws.row_dimensions[row].height = 30; row += 1

    MRT_NAT_ROW = row
    lbl(ws, row, 2, "Кол-во нативных МРТ", "F7F7F7", h="left")
    inp(ws, row, 3, 300)
    for col in [4,5]: ws.cell(row, col).value=None; ws.cell(row, col).fill=fill("FAFAFA"); ws.cell(row, col).border=border()
    ws.cell(row, 6).fill=fill("FFFFFF")
    c=ws.cell(row, 7); c.value="=Нормативы!D8"; c.font=Font(size=9,color="595959",italic=True,name="Calibri"); c.fill=fill("E2EFDA"); c.alignment=align("center"); c.border=border(); c.number_format='0" мин/натив"'
    ws.cell(row, 8).fill=fill("FFFFFF"); ws.cell(row, 8).border=border(); ws.row_dimensions[row].height=20; row+=1

    MRT_SKU_ROW = row
    lbl(ws, row, 2, "Кол-во с контрастом (с КУ)", "F7F7F7", h="left")
    inp(ws, row, 3, 65)
    for col in [4,5]: ws.cell(row, col).value=None; ws.cell(row, col).fill=fill("FAFAFA"); ws.cell(row, col).border=border()
    ws.cell(row, 6).fill=fill("FFFFFF")
    c=ws.cell(row, 7); c.value="=Нормативы!D9"; c.font=Font(size=9,color="595959",italic=True,name="Calibri"); c.fill=fill("E2EFDA"); c.alignment=align("center"); c.border=border(); c.number_format='0" мин/с_КУ"'
    ws.cell(row, 8).fill=fill("FFFFFF"); ws.cell(row, 8).border=border(); ws.row_dimensions[row].height=20; row+=1

    MRT_FAKT_ROW = row
    lbl(ws, row, 2, "ФАКТ (числитель)", "FFF2CC", bold=True, h="left")
    c=ws.cell(row, 3); c.value=f"=C{MRT_NAT_ROW}*Нормативы!D8+C{MRT_SKU_ROW}*Нормативы!D9"; c.font=Font(bold=True,size=10,name="Calibri"); c.fill=fill("FFEB9C"); c.alignment=align("center"); c.border=border_thick(); c.number_format='0" мин"'
    ws.merge_cells(f"D{row}:E{row}"); c2=ws.cell(row,4); c2.value="= натив × 30 мин  +  с_КУ × 45 мин"; c2.font=Font(size=9,italic=True,color="595959",name="Calibri"); c2.fill=fill("FFEB9C"); c2.alignment=align("left"); c2.border=border()
    ws.cell(row, 6).fill=fill("FFFFFF"); ws.cell(row, 7).value="Нат×30 + С_КУ×45 мин"; ws.cell(row, 7).font=Font(size=9,color="595959",name="Calibri"); ws.cell(row, 7).fill=fill("E2EFDA"); ws.cell(row, 7).alignment=align("center"); ws.cell(row, 7).border=border()
    ws.cell(row, 8).fill=fill("FFFFFF"); ws.cell(row, 8).border=border(); ws.row_dimensions[row].height=22; row+=1

    MRT_PLAN_ROW = row
    lbl(ws, row, 2, "ПЛАН (знаменатель)", "E2EFDA", bold=True, h="left")
    c=ws.cell(row, 3); c.value=f"=(Нормативы!D10*Нормативы!D8+Нормативы!D11*Нормативы!D9)*{DAYS_CELL}"; c.font=Font(bold=True,size=10,name="Calibri"); c.fill=fill("C6EFCE"); c.alignment=align("center"); c.border=border_thick(); c.number_format='0" мин"'
    ws.merge_cells(f"D{row}:E{row}"); c2=ws.cell(row,4); c2.value="= (17×30 + 4×45) × рабочих_дней  =  690 × дни"; c2.font=Font(size=9,italic=True,color="595959",name="Calibri"); c2.fill=fill("C6EFCE"); c2.alignment=align("left"); c2.border=border()
    ws.cell(row, 6).fill=fill("FFFFFF"); ws.cell(row, 7).value="690 мин/день × дней"; ws.cell(row, 7).font=Font(size=9,color="595959",name="Calibri"); ws.cell(row, 7).fill=fill("E2EFDA"); ws.cell(row, 7).alignment=align("center"); ws.cell(row, 7).border=border()
    ws.cell(row, 8).fill=fill("FFFFFF"); ws.cell(row, 8).border=border(); ws.row_dimensions[row].height=22; row+=1

    lbl(ws, row, 2, "% ЗАГРУЗКИ МРТ", "1F4E79", bold=True, h="left")
    ws.cell(row, 2).font=Font(bold=True,size=11,color="FFFFFF",name="Calibri")
    c=ws.cell(row, 3); c.value=f"=IF(C{MRT_PLAN_ROW}=0,\"\",C{MRT_FAKT_ROW}/C{MRT_PLAN_ROW}*100)"; c.font=Font(bold=True,size=14,color="1F4E79",name="Calibri"); c.fill=fill("BDD7EE"); c.alignment=align("center"); c.border=border_thick(); c.number_format='0.0"%"'
    ws.merge_cells(f"D{row}:E{row}"); ws.cell(row,4).value="= ФАКТ ÷ ПЛАН × 100"; ws.cell(row,4).font=Font(bold=True,size=10,italic=True,color="1F4E79",name="Calibri"); ws.cell(row,4).fill=fill("BDD7EE"); ws.cell(row,4).alignment=align("left"); ws.cell(row,4).border=border()
    ws.cell(row, 6).fill=fill("FFFFFF")
    sc=ws.cell(row, 7); sc.value=f'=IF(C{row}="","—",IF(C{row}>=85,"✓ Норма","! Ниже нормы"))'; sc.font=Font(bold=True,size=10,name="Calibri"); sc.fill=fill("BDD7EE"); sc.alignment=align("center"); sc.border=border_thick()
    ws.cell(row, 8).fill=fill("FFFFFF"); ws.cell(row, 8).border=border(); ws.row_dimensions[row].height=28; row+=2

    # ══════════════════════════════════════════════════════════════════════════
    # ММГ / РГ / Денситометрия — простые (только штуки)
    # ══════════════════════════════════════════════════════════════════════════
    simple_devices = [
        ("ММГ — Маммография  (взрослые, амбулатория)",    "ММГ",   "Нормативы!D12", "58 исслед./день × дней", "58×дней"),
        ("РГ — Рентгенография  (взрослые, амбулатория)",  "РГ",    "Нормативы!D13", "72 исслед./день × дней", "72×дней"),
        ("Денситометрия  (взрослые, амбулатория)",         "Денс.", "Нормативы!D14", "35 исслед./день × дней", "35×дней"),
    ]
    simple_vals = [1000, 1300, 620]

    for dev_title, dev_short, norm_ref, norm_desc, norm_short, start_val in zip(
            *zip(*simple_devices), simple_vals):
        section_header(row, dev_title)
        row += 1

        for col, txt in zip([2,3,4,5,6,7,8], ["Показатель","Кол-во исследований\n(из витрины)","","","","Норматив","= Результат"]):
            c=ws.cell(row, col); c.value=txt
            c.font=Font(bold=True,size=9,color="FFFFFF" if col==2 else "000000",name="Calibri")
            c.fill=fill("2E75B6" if col==2 else ("FFF2CC" if col==3 else ("E2EFDA" if col==7 else ("DEEAF1" if col==8 else "FFFFFF"))))
            c.alignment=align("center",wrap=True); c.border=border()
        ws.row_dimensions[row].height=30; row+=1

        DEV_FAKT_ROW = row
        lbl(ws, row, 2, "ФАКТ — кол-во исследований", "FFF2CC", bold=True, h="left")
        inp(ws, row, 3, start_val)
        ws.merge_cells(f"D{row}:E{row}"); ws.cell(row,4).value="← кол-во DISTINCT accession_number из витрины"; ws.cell(row,4).font=Font(size=9,italic=True,color="595959",name="Calibri"); ws.cell(row,4).fill=fill("FFFFF0"); ws.cell(row,4).alignment=align("left"); ws.cell(row,4).border=border()
        ws.cell(row, 6).fill=fill("FFFFFF")
        c=ws.cell(row, 7); c.value=norm_ref; c.font=Font(size=9,color="595959",italic=True,name="Calibri"); c.fill=fill("E2EFDA"); c.alignment=align("center"); c.border=border(); c.number_format='0" исслед./день"'
        ws.cell(row, 8).fill=fill("FFFFFF"); ws.cell(row, 8).border=border(); ws.row_dimensions[row].height=22; row+=1

        DEV_PLAN_ROW = row
        lbl(ws, row, 2, "ПЛАН (знаменатель)", "E2EFDA", bold=True, h="left")
        c=ws.cell(row, 3); c.value=f"={norm_ref}*{DAYS_CELL}"; c.font=Font(bold=True,size=10,name="Calibri"); c.fill=fill("C6EFCE"); c.alignment=align("center"); c.border=border_thick(); c.number_format='0" шт"'
        ws.merge_cells(f"D{row}:E{row}"); c2=ws.cell(row,4); c2.value=f"= {norm_short}"; c2.font=Font(size=9,italic=True,color="595959",name="Calibri"); c2.fill=fill("C6EFCE"); c2.alignment=align("left"); c2.border=border()
        ws.cell(row, 6).fill=fill("FFFFFF"); ws.cell(row, 7).value=norm_desc; ws.cell(row, 7).font=Font(size=9,color="595959",name="Calibri"); ws.cell(row, 7).fill=fill("E2EFDA"); ws.cell(row, 7).alignment=align("center"); ws.cell(row, 7).border=border()
        ws.cell(row, 8).fill=fill("FFFFFF"); ws.cell(row, 8).border=border(); ws.row_dimensions[row].height=22; row+=1

        lbl(ws, row, 2, f"% ЗАГРУЗКИ {dev_short}", "1F4E79", bold=True, h="left")
        ws.cell(row, 2).font=Font(bold=True,size=11,color="FFFFFF",name="Calibri")
        c=ws.cell(row, 3); c.value=f"=IF(C{DEV_PLAN_ROW}=0,\"\",C{DEV_FAKT_ROW}/C{DEV_PLAN_ROW}*100)"; c.font=Font(bold=True,size=14,color="1F4E79",name="Calibri"); c.fill=fill("BDD7EE"); c.alignment=align("center"); c.border=border_thick(); c.number_format='0.0"%"'
        ws.merge_cells(f"D{row}:E{row}"); ws.cell(row,4).value="= ФАКТ ÷ ПЛАН × 100"; ws.cell(row,4).font=Font(bold=True,size=10,italic=True,color="1F4E79",name="Calibri"); ws.cell(row,4).fill=fill("BDD7EE"); ws.cell(row,4).alignment=align("left"); ws.cell(row,4).border=border()
        ws.cell(row, 6).fill=fill("FFFFFF")
        sc=ws.cell(row, 7); sc.value=f'=IF(C{row}="","—",IF(C{row}>=85,"✓ Норма","! Ниже нормы"))'; sc.font=Font(bold=True,size=10,name="Calibri"); sc.fill=fill("BDD7EE"); sc.alignment=align("center"); sc.border=border_thick()
        ws.cell(row, 8).fill=fill("FFFFFF"); ws.cell(row, 8).border=border(); ws.row_dimensions[row].height=28; row+=2

    # Легенда внизу
    ws.merge_cells(f"B{row}:H{row}")
    c=ws.cell(row, 2); c.value="Цвета:  Жёлтый = ВВОД ДАННЫХ  |  Зелёный = ПЛАН (знаменатель)  |  Голубой = ФАКТ / РЕЗУЛЬТАТ  |  Нормативы хранятся на листе «Нормативы»"
    c.font=Font(size=9,italic=True,color="595959",name="Calibri"); c.fill=fill("F2F2F2"); c.alignment=align("center"); c.border=border()

    ws.freeze_panes = "B7"
    return ws


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    make_kalkulator(wb)
    make_normativy(wb)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"OK: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
