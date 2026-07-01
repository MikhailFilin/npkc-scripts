"""
export_appeals.py
Выгрузка «Обращения граждан» из CRM Умного процесса Битрикс24 → Excel → PostgreSQL.

Используется как модуль в all_dashboards_up.py (вызов через main()).
VPN не требуется: Битрикс24 API публичен, PostgreSQL в Yandex Cloud доступен без VPN.

INPUT:  Bitrix24 REST API
OUTPUT: appeals_YYYY-MM-DD.xlsx + .csv  (рядом со скриптом)
        PostgreSQL feedback.feedback_2026 (truncate + reload)

Запуск: py -3 export_appeals.py
"""

import logging
import os
import sys
from datetime import date, datetime, timezone

import pandas as pd
import requests
from sqlalchemy import create_engine, text

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPTS_DIR)

import personal_config as _cfg

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Установите openpyxl:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

BITRIX_WEBHOOK = _cfg.BITRIX_WEBHOOK

ENTITY_TYPE_ID = 1172   # «Обращения граждан»
CATEGORY_ID    = 73
PAGE_SIZE      = 50

OUT_FILE = os.path.join(SCRIPTS_DIR, f"appeals_{date.today()}.xlsx")

USER_REF_FIELDS = {"createdBy", "assignedById"}

APPEAL_DATE_FIELD  = "ufCrm65_1761291991550"
APPEAL_DATE_CUTOFF = date(2026, 1, 1)

DATE_ONLY_FIELDS = {APPEAL_DATE_FIELD, "createdTime", "begindate"}
DATETIME_FIELDS  = {"ufCrm65_1764337353608", "ufCrm65_1764337368514"}

DIRECTION_FIELD    = "ufCrm65_1761292641215"
ALLOWED_DIRECTIONS = {"ОМО", "Проведение и описание исследований", "Референс-центр"}
ALLOWED_STAGES     = {"ОМО", "Референс-центр", "ОТВЕТ ОТПРАВЛЕН"}

# ---------------------------------------------------------------------------
# PostgreSQL
# ---------------------------------------------------------------------------

PG_TABLE   = "feedback.feedback_2026"
PG_MATVIEW = "feedback.feedback_all"

# Маппинг: ID поля Bitrix24 → колонка PostgreSQL (не зависит от русских заголовков API)
FIELDS_PG_MAP: dict[str, str] = {
    "id":                        'id',
    "stageId":                   'stadiya',
    "title":                     'nazvanie',
    "createdTime":               'kogda_sozdan',
    "createdBy":                 'kem_sozdan',
    "assignedById":              'otvetstvennyy',
    "begindate":                 'data_nachala',
    "sourceId":                  'istochnik',
    "ufCrm65_1761291991550":     'data_postupleniya_obrashcheniya',
    "ufCrm65_1761292234580":     'otpravit_v_chat_na_opisanie',
    "ufCrm65_1761292392184":     'tip_obrashcheniya',
    "ufCrm65_1761292641215":     'napravlenie',
    "ufCrm65_1761292682780":     'prichina_obrashcheniya',
    "ufCrm65_1761292890455":     'tema_konsultatsii',
    "ufCrm65_1761293136077":     'modalnost_og',
    "ufCrm65_1761293772486":     'vozrastnaya_kategoriya',
    "ufCrm65_1761293834732":     'tekst_obrashcheniya',
    "ufCrm65_1761296760403":     'istochnik_2',
    "ufCrm65_1761298736298":     'status_otveta',
    "ufCrm65_1764337304298":     'mo_mesto_provedeniya_issledovaniya',
    "ufCrm65_1764337353608":     'data_vremya_vypolneniya_issledovaniya',
    "ufCrm65_1764337368514":     'data_vremya_opisaniya_issledovaniya',
    "ufCrm65_1764337382875":     'vremya_opisaniya',
    "ufCrm65_1764337395845":     'zaderzhka_dni',
    "ufCrm65_1764337480040":     'realnaya_prichina_zaderzhki',
    "ufCrm65_1764337557817":     'na_chey_storone_zaderzhka_i_pochemu',
    "ufCrm65_1764337610152":     'realnaya_prichina_peredsmotreniya_issledovaniya',
    "ufCrm65_1764337633":        'fio_vracha',
    "ufCrm65_1764337700179":     'knz_kz_zoh',
    "ufCrm65_1764337713484":     'kommentariy',
    "ufCrm65_1764337738967":     'obektivnost',
    "ufCrm65_1764337763490":     'ispravilos_li_dlya_grazhdanina',
    "ufCrm65_1764337796911":     'tonalnost_obrashcheniya',
    "ufCrm65_1764337864":        'kto_otpravil_otvet',
}

# Фиксированный список колонок в нужном порядке
SELECTED_FIELDS: list[str] = [
    "id",                          # ID
    "stageId",                     # Стадия
    "title",                       # Название
    "createdTime",                 # Когда создан
    "createdBy",                   # Кем создан
    "assignedById",                # Ответственный
    "begindate",                   # Дата начала
    "sourceId",                    # Источник
    "ufCrm65_1761291991550",       # Дата поступления обращения
    "ufCrm65_1761292234580",       # Отправить в чат на описание
    "ufCrm65_1761292392184",       # Тип обращения
    "ufCrm65_1761292641215",       # Направление
    "ufCrm65_1761292682780",       # Причина обращения
    "ufCrm65_1761292890455",       # Тема консультации
    "ufCrm65_1761293136077",       # Модальность (ОГ)
    "ufCrm65_1761293772486",       # Возрастная категория
    "ufCrm65_1761293834732",       # Текст обращения
    "ufCrm65_1761296760403",       # Источник.
    "ufCrm65_1761298736298",       # Статус ответа
    "ufCrm65_1764337304298",       # МО (место проведения исследования)
    "ufCrm65_1764337353608",       # Дата и время выполнения исследования
    "ufCrm65_1764337368514",       # Дата и время описания исследования
    "ufCrm65_1764337382875",       # Время описания
    "ufCrm65_1764337395845",       # Задержка (дни)
    "ufCrm65_1764337480040",       # Реальная причина задержки
    "ufCrm65_1764337557817",       # На чьей стороне задержка и почему?
    "ufCrm65_1764337610152",       # Реальная причина пересмотра исследования
    "ufCrm65_1764337633",          # ФИО врача
    "ufCrm65_1764337700179",       # КНЗ/КЗ/ЗОХ
    "ufCrm65_1764337713484",       # Комментарий
    "ufCrm65_1764337738967",       # Объективность
    "ufCrm65_1764337763490",       # Исправилось ли для гражданина
    "ufCrm65_1764337796911",       # Тональность обращения
    "ufCrm65_1764337864",          # Кто отправил ответ
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger(__name__)

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FF")


# ---------------------------------------------------------------------------
# Bitrix24 helpers
# ---------------------------------------------------------------------------

def bx(method: str, params: dict | None = None) -> dict:
    url = f"{BITRIX_WEBHOOK}/{method}.json"
    try:
        r = requests.post(url, json=params or {}, timeout=30)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        log.error("API [%s]: %s", method, e)
        return {}


def fetch_all_items(entity_type_id: int, category_id: int | None) -> list[dict]:
    items: list[dict] = []
    start = 0
    params: dict = {"entityTypeId": entity_type_id}
    if category_id is not None:
        params["filter"] = {"categoryId": category_id}
    while True:
        params["start"] = start
        resp = bx("crm.item.list", params)
        result = resp.get("result", {})
        batch = result.get("items", []) if isinstance(result, dict) else []
        if not batch:
            break
        items.extend(batch)
        total = resp.get("total", 0)
        log.info("Загружено: %d / %d", len(items), total)
        if len(items) >= total:
            break
        start += PAGE_SIZE
    return items


def fetch_field_defs(
    entity_type_id: int,
) -> tuple[dict[str, str], dict[str, str], dict[str, str], dict[str, dict[str, str]], set[str], set[str]]:
    resp = bx("crm.item.fields", {"entityTypeId": entity_type_id})
    fields = resp.get("result", {}).get("fields", {})
    label_map: dict[str, str] = {}
    enum_map: dict[str, dict[str, str]] = {}
    user_fields: set[str] = set()
    multi_fields: set[str] = set()
    for name, fdef in fields.items():
        label_map[name] = fdef.get("title", name)
        if fdef.get("type") in ("user", "employee"):
            user_fields.add(name)
        if fdef.get("type") == "enumeration":
            enum_map[name] = {
                str(item["ID"]): item["VALUE"] for item in fdef.get("items", [])
            }
        if fdef.get("isMultiple"):
            multi_fields.add(name)

    status_resp = bx("crm.status.list")
    statuses = status_resp.get("result", [])

    stage_map: dict[str, str] = {}
    stage_prefix = f"DYNAMIC_{entity_type_id}_STAGE_"
    for s in statuses:
        if str(s.get("ENTITY_ID", "")).startswith(stage_prefix):
            stage_map[s["STATUS_ID"]] = s["NAME"]

    source_map: dict[str, str] = {
        s["STATUS_ID"]: s["NAME"] for s in statuses if s.get("ENTITY_ID") == "SOURCE"
    }

    return label_map, stage_map, source_map, enum_map, user_fields, multi_fields


def resolve_users(user_ids: set[int]) -> dict[int, str]:
    if not user_ids:
        return {}
    names: dict[int, str] = {}
    ids = list(user_ids)
    for i in range(0, len(ids), 50):
        batch = ids[i : i + 50]
        resp = bx("user.get", {"FILTER": {"ID": batch}})
        for u in resp.get("result", []):
            uid = int(u["ID"])
            full = " ".join(
                filter(None, [u.get("LAST_NAME"), u.get("NAME"), u.get("SECOND_NAME")])
            )
            names[uid] = full.strip() or str(uid)
    return names


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _cell_val(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Да" if v else "Нет"
    if isinstance(v, list):
        return "; ".join(str(x) for x in v if x is not None)
    return str(v)


def _parse_bx_datetime(raw):
    if not raw:
        return None
    try:
        return datetime.fromisoformat(str(raw))
    except ValueError:
        return None


def _format_bx_date(raw, *, date_only: bool):
    dt = _parse_bx_datetime(raw)
    if dt is None:
        return raw
    if date_only:
        return dt.strftime("%d.%m.%Y")
    return dt.strftime("%d.%m.%Y %H.%M.%S")


def _resolve_value(
    key: str,
    raw,
    *,
    stage_map: dict[str, str],
    source_map: dict[str, str],
    enum_map: dict[str, dict[str, str]],
    user_fields: set[str],
    multi_fields: set[str],
    user_names: dict[int, str],
):
    if raw is None:
        return raw
    if key in DATE_ONLY_FIELDS:
        return _format_bx_date(raw, date_only=True)
    if key in DATETIME_FIELDS:
        return _format_bx_date(raw, date_only=False)
    if key in user_fields:
        values = raw if isinstance(raw, list) else [raw]
        resolved = []
        for v in values:
            try:
                resolved.append(user_names.get(int(v), str(v)))
            except (ValueError, TypeError):
                resolved.append(str(v))
        return resolved if key in multi_fields else (resolved[0] if resolved else None)
    if key == "stageId":
        return stage_map.get(str(raw), str(raw))
    if key == "sourceId":
        return source_map.get(str(raw), str(raw))
    if key in enum_map:
        values = raw if isinstance(raw, list) else [raw]
        resolved = [enum_map[key].get(str(v), str(v)) for v in values]
        return resolved if key in multi_fields else (resolved[0] if resolved else None)
    return raw


def save_xlsx(
    items: list[dict],
    label_map: dict[str, str],
    stage_map: dict[str, str],
    source_map: dict[str, str],
    enum_map: dict[str, dict[str, str]],
    user_fields: set[str],
    multi_fields: set[str],
    user_names: dict[int, str],
) -> None:
    if not items:
        log.warning("Нет данных — файл не создан.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Обращения граждан"
    ws.freeze_panes = "A2"

    for col, key in enumerate(SELECTED_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=label_map.get(key, key))
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for row_idx, item in enumerate(items, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col, key in enumerate(SELECTED_FIELDS, 1):
            raw = _resolve_value(
                key, item.get(key),
                stage_map=stage_map, source_map=source_map, enum_map=enum_map,
                user_fields=user_fields, multi_fields=multi_fields, user_names=user_names,
            )
            c = ws.cell(row=row_idx, column=col, value=_cell_val(raw))
            if fill:
                c.fill = fill

    for col, key in enumerate(SELECTED_FIELDS, 1):
        label   = label_map.get(key, key)
        max_len = max(len(label), *(len(_cell_val(item.get(key, ""))) for item in items))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)

    ws.row_dimensions[1].height = 30
    wb.save(OUT_FILE)
    log.info("Готово: %s  (%d строк)", OUT_FILE, len(items))


def export_to_csv(xlsx_path: str) -> tuple[str, pd.DataFrame]:
    df = pd.read_excel(xlsx_path)
    df = df[pd.to_numeric(df["ID"], errors="coerce").notna()]
    df = df.fillna("Не заполнено")
    df = df[~df["Тип обращения"].str.contains("Повтор", na=False)]

    df["Когда создан"] = pd.to_datetime(df["Когда создан"], format="%d.%m.%Y", utc=True)
    df["first_day_of_week"] = (
        df["Когда создан"].dt.to_period("W").dt.start_time.astype(str)
    )
    df["now"] = datetime.now(timezone.utc).isoformat()

    csv_path = os.path.splitext(xlsx_path)[0] + ".csv"
    df.to_csv(csv_path, sep=";", index=False)
    log.info("CSV готов: %s  (%d строк)", csv_path, len(df))
    return csv_path, df


# ---------------------------------------------------------------------------
# PostgreSQL upload
# ---------------------------------------------------------------------------

def build_pg_df(
    items: list[dict],
    stage_map: dict[str, str],
    source_map: dict[str, str],
    enum_map: dict[str, dict[str, str]],
    user_fields: set[str],
    multi_fields: set[str],
    user_names: dict[int, str],
) -> pd.DataFrame:
    """Строит DataFrame для PG напрямую из items по FIELDS_PG_MAP (без зависимости от русских заголовков)."""
    rows = []
    for item in items:
        row: dict = {}
        for bx_key, pg_col in FIELDS_PG_MAP.items():
            raw = _resolve_value(
                bx_key, item.get(bx_key),
                stage_map=stage_map, source_map=source_map, enum_map=enum_map,
                user_fields=user_fields, multi_fields=multi_fields, user_names=user_names,
            )
            row[pg_col] = _cell_val(raw) if raw is not None else "Не заполнено"
        rows.append(row)

    df = pd.DataFrame(rows).fillna("Не заполнено")

    # Фильтр «Повтор» — как в export_to_csv
    df = df[~df['tip_obrashcheniya'].str.contains("Повтор", na=False)]

    # Вычисляемые колонки
    try:
        df['first_day_of_week'] = (
            pd.to_datetime(df['kogda_sozdan'], format="%d.%m.%Y", utc=True, errors='coerce')
            .dt.to_period("W").dt.start_time.astype(str)
        )
    except Exception:
        df['first_day_of_week'] = ""
    df['now'] = datetime.now(timezone.utc).isoformat()

    return df


def _log_empty_cols(df: pd.DataFrame) -> None:
    """Логирует колонки, у которых > 80% пустых значений — для диагностики."""
    for col in df.columns:
        empty = ((df[col] == "") | df[col].isna()).sum()
        pct = 100 * empty / len(df) if len(df) else 0
        if pct > 80:
            log.warning("⚠️  Колонка '%s': %d/%d пустых (%.0f%%)", col, empty, len(df), pct)
        else:
            log.debug("   Колонка '%s': %d/%d пустых (%.0f%%)", col, empty, len(df), pct)


def upload_to_postgres(df: pd.DataFrame) -> None:
    """Очищает PG_TABLE и загружает df (колонки уже в PG-формате из build_pg_df)."""
    log.info("Диагностика заполненности колонок:")
    _log_empty_cols(df)

    pg_url = (
        f"postgresql://{_cfg.PG_USER}:{_cfg.PG_PASSWORD}"
        f"@{_cfg.PG_HOST}:{_cfg.PG_PORT}/{_cfg.PG_DATABASE}"
    )
    engine = create_engine(pg_url)
    schema, table = PG_TABLE.split(".")

    log.info("PostgreSQL: подключаемся к %s …", _cfg.PG_HOST)
    try:
        with engine.begin() as conn:
            conn.execute(text(f"TRUNCATE TABLE {PG_TABLE}"))
        log.info("Таблица %s очищена", PG_TABLE)

        df.to_sql(
            table, engine, schema=schema,
            if_exists="append", index=False,
            method="multi", chunksize=500,
        )
        log.info("Загружено %d строк в %s", len(df), PG_TABLE)

        try:
            with engine.begin() as conn:
                conn.execute(text(f"REFRESH MATERIALIZED VIEW {PG_MATVIEW}"))
            log.info("Материализованное представление %s обновлено", PG_MATVIEW)
        except Exception as view_exc:
            log.info(
                "%s является обычным VIEW и обновляется автоматически — REFRESH не требуется (%s)",
                PG_MATVIEW, view_exc.__class__.__name__,
            )
    except Exception as exc:
        log.error("Ошибка PostgreSQL: %s", exc)
        raise
    finally:
        engine.dispose()


# ---------------------------------------------------------------------------
# Main (возвращает True/False для all_dashboards_up)
# ---------------------------------------------------------------------------

def main() -> bool:
    log.info("SPA entityTypeId = %d, categoryId = %d", ENTITY_TYPE_ID, CATEGORY_ID)

    log.info("Получаем описание полей …")
    label_map, stage_map, source_map, enum_map, user_fields, multi_fields = fetch_field_defs(
        ENTITY_TYPE_ID
    )
    if not label_map:
        log.error("Не удалось получить поля. Проверьте entityTypeId и права вебхука (scope crm).")
        return False

    log.info("Полей: %d, стадий: %d, источников: %d", len(label_map), len(stage_map), len(source_map))

    log.info("Загружаем элементы …")
    items = fetch_all_items(ENTITY_TYPE_ID, CATEGORY_ID)
    if not items:
        log.warning("Элементов не найдено.")
        return False
    log.info("Итого: %d записей", len(items))

    def _appeal_date_ok(item: dict) -> bool:
        dt = _parse_bx_datetime(item.get(APPEAL_DATE_FIELD))
        return dt is not None and dt.date() >= APPEAL_DATE_CUTOFF

    items = [item for item in items if _appeal_date_ok(item)]
    log.info(
        "После фильтра по дате поступления обращения (с %s): %d записей",
        APPEAL_DATE_CUTOFF.strftime("%d.%m.%Y"), len(items),
    )
    if not items:
        log.warning("После фильтра по дате не осталось записей.")
        return False

    def _direction_stage_ok(item: dict) -> bool:
        stage_label    = stage_map.get(str(item.get("stageId")), str(item.get("stageId")))
        raw_direction  = item.get(DIRECTION_FIELD)
        direction_label = enum_map.get(DIRECTION_FIELD, {}).get(str(raw_direction), raw_direction)
        return stage_label in ALLOWED_STAGES and direction_label in ALLOWED_DIRECTIONS

    items = [item for item in items if _direction_stage_ok(item)]
    log.info("После фильтра по направлению и стадии: %d записей", len(items))
    if not items:
        log.warning("После фильтра по направлению/стадии не осталось записей.")
        return False

    uid_set: set[int] = set()
    for item in items:
        for uf in user_fields:
            v = item.get(uf)
            values = v if isinstance(v, list) else [v]
            for val in values:
                if val is not None:
                    try:
                        uid_set.add(int(val))
                    except (ValueError, TypeError):
                        pass

    if uid_set:
        log.info("Разрешаем %d пользователей …", len(uid_set))
    user_names = resolve_users(uid_set)

    save_xlsx(items, label_map, stage_map, source_map, enum_map, user_fields, multi_fields, user_names)
    export_to_csv(OUT_FILE)

    df_pg = build_pg_df(items, stage_map, source_map, enum_map, user_fields, multi_fields, user_names)
    log.info("Строк для PG (после фильтра Повтор): %d", len(df_pg))
    upload_to_postgres(df_pg)
    return True


if __name__ == "__main__":
    ok = main()
    sys.exit(0 if ok else 1)
